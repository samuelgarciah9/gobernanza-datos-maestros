"""Genera los Excel de captura por figura desde la vista de merge (buckets 1 y 3).

Portado de 08_export_decisiones.py a función importable para la app Qt.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from gd.conexion import BASE, get_connection

VISTA_MERGE = "V_GD_MERGE_MATERIALES"
TABLA_CAT = "GD_CAT_RAZONES"
ESTADOS = ("MIGRAR", "DESCARTAR", "PENDIENTE")
BUCKETS_CAPTURA = ("POR_DECIDIR", "RE_REVISAR")
SEGMENTOS = ("PRODUCTIVO", "NO_PRODUCTIVO")

COLS_CONTEXTO = [
    "NUMERO_PRODUCTO_ANTIGUO", "TIPO_MATERIAL", "BUCKET", "MOTIVO", "PT_PROD_LINE",
    "DESCRIPCION", "GRUPO_PRODUCTO", "PT_GROUP", "PT_PART_TYPE", "EXIST_TOTAL",
    "ULTIMA_TRANSACCION", "ULTIMO_TIPO_TXN", "DIAS_SIN_MOVIMIENTO",
    "HAS_EXIST", "HAS_PO", "HAS_SO", "HAS_WO", "HAS_INV_SEG", "HAS_DIST", "HAS_RET",
    "PESO_ACTIVIDAD", "DECISION_PREVIA", "FLAGS_CAMBIADOS",
]
COLS_BANDERA = {"HAS_EXIST", "HAS_PO", "HAS_SO", "HAS_WO", "HAS_INV_SEG", "HAS_DIST", "HAS_RET"}
DERIVADAS = {"TIPO_MATERIAL"}
SENALES_FIRMES = ("HAS_EXIST", "HAS_PO", "HAS_SO", "HAS_WO", "HAS_INV_SEG")
COL_SUGERENCIA = "ESTADO_SUGERIDO"
COLS_DECISION = ["ESTADO", "RAZON", "COMENTARIO"]
COLS_CONTROL = ["PT_DOMAIN", "RUN_ID_AL_DECIDIR", "HASH_AL_DECIDIR"]

AZUL, GRIS, AMARILLO, VERDE = "1F4E78", "808080", "BF8F00", "375623"
BORDE = Border(*(Side(style="thin", color="D9D9D9"),) * 4)


def tipo_material(fila: dict) -> str:
    pl = (fila.get("PT_PROD_LINE") or "").strip().upper()
    return "NO_PRODUCTIVO" if pl == "REF" else "PRODUCTIVO"


def sugerir_estado(fila: dict) -> str:
    if any(fila.get(f) for f in SENALES_FIRMES):
        return "MIGRAR"
    if fila.get("HAS_DIST") or fila.get("HAS_RET") or fila.get("HAS_ULT_TXN"):
        return "PENDIENTE"
    return "DESCARTAR"


def fetch_datos(dominio: str):
    with get_connection() as con:
        cur = con.cursor()
        cur.execute(f"SELECT RAZON FROM {TABLA_CAT} WHERE ACTIVO='S' ORDER BY ESTADO_SUGERIDO, RAZON")
        razones = [r[0] for r in cur.fetchall()]
        reales = [c for c in COLS_CONTEXTO if c not in DERIVADAS]
        necesarias = list(dict.fromkeys(
            reales + ["PT_PROD_LINE", "HAS_ULT_TXN", "BANDERAS_HASH", "PT_DOMAIN", "RUN_ID"]))
        marcadores = ", ".join(f"'{b}'" for b in BUCKETS_CAPTURA)
        cur.execute(
            f"SELECT {', '.join(necesarias)} FROM {VISTA_MERGE} "
            f"WHERE PT_DOMAIN = :d AND BUCKET IN ({marcadores}) "
            f"ORDER BY CASE WHEN BUCKET='RE_REVISAR' THEN 0 ELSE 1 END, NUMERO_PRODUCTO_ANTIGUO",
            d=dominio)
        cols = [c[0] for c in cur.description]
        filas = [dict(zip(cols, vals)) for vals in cur.fetchall()]
    for f in filas:
        f["TIPO_MATERIAL"] = tipo_material(f)
    run_id = filas[0]["RUN_ID"] if filas else None
    return run_id, razones, filas


def construir_excel(dominio, tipo, run_id, razones, filas, destino: Path | None = None) -> Path:
    carpeta = Path(destino) if destino else (BASE / "entregables")
    salida = carpeta / f"decisiones_{dominio}_{tipo}_{run_id}_para_captura.xlsx"
    wb = Workbook()

    ins = wb.active
    ins.title = "Instrucciones"
    for i, (txt, bold) in enumerate([
        (f"Captura de decisiones de migracion  -  Dominio {dominio}  -  {tipo}", True),
        ("", False),
        (f"Foto (snapshot): {run_id}   |   Materiales: {len(filas)}", False),
        ("", False),
        ("Cada fila es un material. Decida si se MIGRA o se DESCARTA (o queda PENDIENTE).", False),
        ("Las columnas grises son referencia (no se editan).", False),
        ("", False),
        ("Solo llene las columnas verdes:", True),
        ("  - ESTADO:  elija de la lista -> MIGRAR / DESCARTAR / PENDIENTE.", False),
        ("  - RAZON:  elija una razon del catalogo (obligatoria si ESTADO no es PENDIENTE).", False),
        ("  - COMENTARIO:  nota libre opcional.", False),
        ("", False),
        ("BUCKET: POR_DECIDIR = nuevo/sin decidir; RE_REVISAR = cambio algo (ver MOTIVO y", False),
        ("FLAGS_CAMBIADOS) sobre una decision previa: reconsidere.", False),
    ], start=1):
        c = ins.cell(row=i, column=1, value=txt)
        c.font = Font(bold=bold, size=12 if (bold and i == 1) else 11, color=AZUL if i == 1 else "000000")
    ins.column_dimensions["A"].width = 95

    cat = wb.create_sheet("Catalogo")
    cat["A1"] = "RAZONES (no editar)"
    cat["A1"].font = Font(bold=True)
    for i, r in enumerate(razones, start=2):
        cat[f"A{i}"] = r
    cat.column_dimensions["A"].width = 60
    cat.sheet_state = "hidden"

    ws = wb.create_sheet("Decisiones")
    todos = COLS_CONTEXTO + [COL_SUGERENCIA] + COLS_DECISION + COLS_CONTROL
    for j, nombre in enumerate(todos, start=1):
        c = ws.cell(row=1, column=j, value=nombre)
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        fill = AMARILLO if nombre == COL_SUGERENCIA else (VERDE if nombre in COLS_DECISION else GRIS)
        c.fill = PatternFill("solid", fgColor=fill)
        c.border = BORDE

    for i, fila in enumerate(filas, start=2):
        sug = sugerir_estado(fila)
        for j, nombre in enumerate(todos, start=1):
            if nombre in COLS_DECISION:
                valor = None
            elif nombre == COL_SUGERENCIA:
                valor = sug
            elif nombre == "RUN_ID_AL_DECIDIR":
                valor = run_id
            elif nombre == "HASH_AL_DECIDIR":
                valor = fila.get("BANDERAS_HASH")
            elif nombre in COLS_BANDERA:
                valor = "Si" if fila.get(nombre) else "No"
            else:
                valor = fila.get(nombre)
            c = ws.cell(row=i, column=j, value=valor)
            c.border = BORDE
            c.protection = Protection(locked=nombre not in COLS_DECISION)

    n_filas = len(filas) + 1
    letra_estado = get_column_letter(todos.index("ESTADO") + 1)
    dv_e = DataValidation(type="list", formula1=f'"{",".join(ESTADOS)}"', allow_blank=True)
    dv_e.error = "Elija un estado de la lista."
    ws.add_data_validation(dv_e)
    dv_e.add(f"{letra_estado}2:{letra_estado}{n_filas}")

    letra_razon = get_column_letter(todos.index("RAZON") + 1)
    dv_r = DataValidation(type="list", formula1=f"Catalogo!$A$2:$A${len(razones) + 1}", allow_blank=True)
    dv_r.error = "Elija una razon del catalogo."
    ws.add_data_validation(dv_r)
    dv_r.add(f"{letra_razon}2:{letra_razon}{n_filas}")

    anchos = {
        "NUMERO_PRODUCTO_ANTIGUO": 22, "TIPO_MATERIAL": 15, "PT_PROD_LINE": 13,
        "BUCKET": 13, "MOTIVO": 18, "DECISION_PREVIA": 15, "FLAGS_CAMBIADOS": 22,
        "DESCRIPCION": 34, "GRUPO_PRODUCTO": 20, "PT_GROUP": 16, "PT_PART_TYPE": 16,
        "ULTIMA_TRANSACCION": 16, "ULTIMO_TIPO_TXN": 16, "DIAS_SIN_MOVIMIENTO": 14,
        "PESO_ACTIVIDAD": 14, "ESTADO_SUGERIDO": 15, "ESTADO": 14, "RAZON": 42,
        "COMENTARIO": 30, "RUN_ID_AL_DECIDIR": 28, "HASH_AL_DECIDIR": 14,
    }
    for j, nombre in enumerate(todos, start=1):
        ws.column_dimensions[get_column_letter(j)].width = anchos.get(nombre, 11)
    for nombre in COLS_CONTROL:
        ws.column_dimensions[get_column_letter(todos.index(nombre) + 1)].hidden = True

    ws.freeze_panes = ws.cell(row=2, column=len(COLS_CONTEXTO) + 1)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(todos))}{n_filas}"
    ws.protection.sheet = True
    ws.protection.autoFilter = False
    ws.protection.sort = False

    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(salida)
    return salida


def generar(dominios=("ST", "RSS"), destino: Path | None = None) -> list[str]:
    """Genera los Excel por figura de cada dominio. Devuelve mensajes de resultado.

    destino: carpeta donde guardar los Excel. Si es None, usa 'entregables'.
    """
    msgs = []
    for dominio in dominios:
        run_id, razones, filas = fetch_datos(dominio)
        for seg in SEGMENTOS:
            sub = [f for f in filas if f["TIPO_MATERIAL"] == seg]
            if not sub:
                continue
            ruta = construir_excel(dominio, seg, run_id, razones, sub, destino)
            msgs.append(f"{len(sub)} materiales {dominio}/{seg} -> {ruta.name}")
    return msgs
