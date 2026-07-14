"""Lectura del Excel de captura y UPSERT de decisiones — portado del importador."""

from __future__ import annotations

import json

from openpyxl import load_workbook

from gd.conexion import BASE, get_connection

TABLA = "GD_DECISIONES_MATERIALES"
ESTADOS_VALIDOS = {"MIGRAR", "DESCARTAR", "PENDIENTE"}
ROLES = ["COMPRAS", "INVENTARIOS", "INGENIERIA"]
CONFIG = BASE / ".importador_config.json"

HOJA = "Decisiones"
COL_MAT, COL_DOM = "NUMERO_PRODUCTO_ANTIGUO", "PT_DOMAIN"
COL_ESTADO, COL_RAZON, COL_COMENT = "ESTADO", "RAZON", "COMENTARIO"
COL_RUN, COL_HASH = "RUN_ID_AL_DECIDIR", "HASH_AL_DECIDIR"


def leer_decisiones(ruta):
    wb = load_workbook(ruta, data_only=True)
    if HOJA not in wb.sheetnames:
        raise ValueError(f"El archivo no tiene la hoja '{HOJA}'. ¿Es el Excel de captura correcto?")
    ws = wb[HOJA]
    hdr = {}
    for idx, cell in enumerate(ws[1]):
        if cell.value is not None:
            hdr[str(cell.value).strip()] = idx
    faltan = [c for c in (COL_MAT, COL_ESTADO, COL_RAZON, COL_RUN, COL_HASH) if c not in hdr]
    if faltan:
        raise ValueError(f"Al Excel le faltan columnas requeridas: {', '.join(faltan)}")
    tiene_dom, tiene_coment = COL_DOM in hdr, COL_COMENT in hdr

    def val(row, col):
        v = row[hdr[col]].value if col in hdr else None
        if v is None:
            return None
        v = str(v).strip()
        return v or None

    decisiones, errores, sin_decidir = [], [], 0
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if all(c.value in (None, "") for c in row):
            continue
        estado = val(row, COL_ESTADO)
        if not estado:
            sin_decidir += 1
            continue
        estado = estado.upper()
        mat, razon = val(row, COL_MAT), val(row, COL_RAZON)
        if mat:
            mat = mat.upper()  # llave canonica en MAYUSCULAS (para que cruce con el snapshot)
        coment = val(row, COL_COMENT) if tiene_coment else None
        run, hsh = val(row, COL_RUN), val(row, COL_HASH)
        dom = val(row, COL_DOM) if tiene_dom else None
        if not dom and run:
            partes = run.split("_")
            dom = partes[1].upper() if len(partes) >= 2 else None
        dom = dom.upper() if dom else None
        etq = f"Fila {i}" + (f" (material {mat})" if mat else "")
        if estado not in ESTADOS_VALIDOS:
            errores.append(f"{etq}: ESTADO inválido '{estado}'.")
            continue
        if estado != "PENDIENTE" and not razon:
            errores.append(f"{etq}: falta RAZÓN (obligatoria si ESTADO no es PENDIENTE).")
            continue
        if not (mat and dom and run and hsh):
            errores.append(f"{etq}: faltan datos de control (material/dominio/run/hash).")
            continue
        decisiones.append({"mat": mat, "dom": dom, "estado": estado, "razon": razon,
                           "coment": coment, "run": run, "hash": hsh})
    return decisiones, errores, sin_decidir


_MERGE = f"""
MERGE INTO {TABLA} d
USING (SELECT :mat AS NUMERO_PRODUCTO_ANTIGUO, :dom AS PT_DOMAIN FROM dual) s
   ON (d.NUMERO_PRODUCTO_ANTIGUO = s.NUMERO_PRODUCTO_ANTIGUO AND d.PT_DOMAIN = s.PT_DOMAIN)
WHEN MATCHED THEN UPDATE SET
     ESTADO=:estado, RAZON=:razon, COMENTARIO=:coment, DECIDIDO_POR=:quien, ROL=:rol,
     FECHA_DECISION=SYSDATE, HASH_AL_DECIDIR=:hash, RUN_ID_AL_DECIDIR=:run
WHEN NOT MATCHED THEN INSERT
     (NUMERO_PRODUCTO_ANTIGUO, PT_DOMAIN, ESTADO, RAZON, COMENTARIO, DECIDIDO_POR, ROL,
      FECHA_DECISION, HASH_AL_DECIDIR, RUN_ID_AL_DECIDIR)
     VALUES (:mat, :dom, :estado, :razon, :coment, :quien, :rol, SYSDATE, :hash, :run)
"""


def guardar_decisiones(decisiones, quien, rol):
    with get_connection() as con:
        cur = con.cursor()
        cur.execute(f"SELECT NUMERO_PRODUCTO_ANTIGUO, PT_DOMAIN FROM {TABLA}")
        existentes = {(m, dm) for m, dm in cur.fetchall()}
        nuevas = sum(1 for x in decisiones if (x["mat"], x["dom"]) not in existentes)
        cur.executemany(_MERGE, [dict(x, quien=quien[:60], rol=rol) for x in decisiones])
        con.commit()
    return nuevas, len(decisiones) - nuevas


def cargar_config() -> dict:
    try:
        return json.loads(CONFIG.read_text(encoding="utf-8"))
    except Exception:
        return {}


def guardar_config(nombre, rol):
    try:
        CONFIG.write_text(json.dumps({"nombre": nombre, "rol": rol}), encoding="utf-8")
    except Exception:
        pass
