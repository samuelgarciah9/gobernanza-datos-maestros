"""Genera los Excel de captura de PROVEEDORES por sociedad (buckets 1 y 3).

Espejo de gd/exportador.py. Un Excel por sociedad (ST/RSS) con los proveedores
POR_DECIDIR y RE_REVISAR. Llave = RFC. La figura llena ESTADO/RAZON/COMENTARIO.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from gd.conexion import BASE, get_connection

VISTA_MERGE = "V_GD_PROV_MERGE"
ESTADOS = ("MIGRAR", "DESCARTAR", "PENDIENTE")
BUCKETS_CAPTURA = ("POR_DECIDIR", "RE_REVISAR")

# Catálogo de razones para proveedores (dropdown en el Excel).
RAZONES = [
    "Con compras en 2025 - se migra",
    "Con orden viva / recepcion / pago - se migra",
    "Proveedor estrategico - se migra",
    "Sin actividad - obsoleto",
    "Duplicado (mismo RFC) - se decide por RFC",
    "Datos incompletos - pendiente de revision",
]

COLS_CONTEXTO = [
    "RFC", "NOMBRE_CONSOLIDADO", "TIPO_PROVEEDOR", "PERSONA_FISICA", "BUCKET", "MOTIVO",
    "PESO_ACTIVIDAD", "HAS_BACKORDER", "HAS_RECEPT", "HAS_PENDING", "HAS_VOUCHER",
    "OC_2025", "FECHA_ULT_COMPRA", "MONTO_MXN_2025", "MONTO_USD_2025",
    "OC_VIVAS", "MONTO_PEND", "CANTIDAD_VOUCHERS", "DECISION_PREVIA", "FLAGS_CAMBIADOS",
]
COLS_BANDERA = {"HAS_BACKORDER", "HAS_RECEPT", "HAS_PENDING", "HAS_VOUCHER"}
COL_SUGERENCIA = "ESTADO_SUGERIDO"
COLS_DECISION = ["ESTADO", "RAZON", "COMENTARIO"]
COLS_CONTROL = ["DOMINIO", "RUN_ID_AL_DECIDIR", "HASH_AL_DECIDIR"]

AZUL, GRIS, AMARILLO, VERDE = "1F4E78", "808080", "BF8F00", "375623"
BORDE = Border(*(Side(style="thin", color="D9D9D9"),) * 4)


def sugerir_estado(fila: dict) -> str:
    return "MIGRAR" if (fila.get("PESO_ACTIVIDAD") or 0) > 0 else "DESCARTAR"


def fetch_datos(dominio: str):
    with get_connection() as con:
        cur = con.cursor()
        necesarias = list(dict.fromkeys(
            COLS_CONTEXTO + ["BANDERAS_HASH", "DOMINIO", "RUN_ID"]))
        # 'RUN_ID' y 'BANDERAS_HASH' vienen del snapshot dentro del merge
        marcadores = ", ".join(f"'{b}'" for b in BUCKETS_CAPTURA)
        cur.execute(
            f"SELECT {', '.join(necesarias)} FROM {VISTA_MERGE} "
            f"WHERE DOMINIO = :d AND BUCKET IN ({marcadores}) "
            f"ORDER BY CASE WHEN BUCKET='RE_REVISAR' THEN 0 ELSE 1 END, "
            f"        CASE WHEN PESO_ACTIVIDAD > 0 THEN 0 ELSE 1 END, NOMBRE_CONSOLIDADO",
            d=dominio)
        cols = [c[0] for c in cur.description]
        filas = [dict(zip(cols, vals)) for vals in cur.fetchall()]
    run_id = filas[0]["RUN_ID"] if filas else None
    return run_id, filas


def construir_excel(dominio, run_id, filas, destino: Path | None = None) -> Path:
    carpeta = Path(destino) if destino else (BASE / "entregables")
    salida = carpeta / f"decisiones_PROV_{dominio}_{run_id}_para_captura.xlsx"
    wb = Workbook()

    ins = wb.active
    ins.title = "Instrucciones"
    for i, (txt, bold) in enumerate([
        (f"Captura de decisiones de proveedores  -  Sociedad {dominio}", True),
        ("", False),
        (f"Foto (snapshot): {run_id}   |   Proveedores: {len(filas)}", False),
        ("", False),
        ("Cada fila es un proveedor (por RFC). Decida si se MIGRA o se DESCARTA (o PENDIENTE).", False),
        ("Las columnas grises son referencia (no se editan).", False),
        ("", False),
        ("Solo llene las columnas verdes:", True),
        ("  - ESTADO:  elija de la lista -> MIGRAR / DESCARTAR / PENDIENTE.", False),
        ("  - RAZON:  elija una razon del catalogo (obligatoria si ESTADO no es PENDIENTE).", False),
        ("  - COMENTARIO:  nota libre opcional.", False),
        ("", False),
        ("Nota: los proveedores que YA son BP en SAP no aparecen aqui (ya quedaron como", False),
        ("MIGRAR; solo falta extenderlos a la sociedad FI 1200 de ST).", False),
    ], start=1):
        c = ins.cell(row=i, column=1, value=txt)
        c.font = Font(bold=bold, size=12 if (bold and i == 1) else 11, color=AZUL if i == 1 else "000000")
    ins.column_dimensions["A"].width = 95

    cat = wb.create_sheet("Catalogo")
    cat["A1"] = "RAZONES (no editar)"
    cat["A1"].font = Font(bold=True)
    for i, r in enumerate(RAZONES, start=2):
        cat[f"A{i}"] = r
    cat.column_dimensions["A"].width = 60
    cat.sheet_state = "hidden"

    ws = wb.create_sheet("Decisiones")
    todos = COLS_CONTEXTO + [COL_SUGERENCIA] + COLS_DECISION + COLS_CONTROL
    for j, nombre in enumerate(todos, start=1):
        c = ws.cell(row=1, column=j, value=nombre)
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        fill = AMARILLO if nombre == COL_SUGERENCIA else (VERDE if nombre in COLS_DECISION else GRIS)
        c.fill = PatternFill("solid", fgColor=fill)
        c.border = BORDE

    for i, fila in enumerate(filas, start=2):
        sug = sugerir_estado(fila)
        for j, nombre in enumerate(todos, start=1):
            if nombre in COLS_DECISION:
                valor = None
            elif nombre == COL_SUGERENCIA:
                valor = sug
            elif nombre == "RUN_ID_AL_DECIDIR":
                valor = run_id
            elif nombre == "HASH_AL_DECIDIR":
                valor = fila.get("BANDERAS_HASH")
            elif nombre in COLS_BANDERA:
                valor = "Si" if fila.get(nombre) else "No"
            else:
                valor = fila.get(nombre)
            c = ws.cell(row=i, column=j, value=valor)
            c.border = BORDE
            c.protection = Protection(locked=nombre not in COLS_DECISION)

    n_filas = len(filas) + 1
    letra_estado = get_column_letter(todos.index("ESTADO") + 1)
    dv_e = DataValidation(type="list", formula1=f'"{",".join(ESTADOS)}"', allow_blank=True)
    dv_e.error = "Elija un estado de la lista."
    ws.add_data_validation(dv_e)
    dv_e.add(f"{letra_estado}2:{letra_estado}{n_filas}")

    letra_razon = get_column_letter(todos.index("RAZON") + 1)
    dv_r = DataValidation(type="list", formula1=f"Catalogo!$A$2:$A${len(RAZONES) + 1}", allow_blank=True)
    dv_r.error = "Elija una razon del catalogo."
    ws.add_data_validation(dv_r)
    dv_r.add(f"{letra_razon}2:{letra_razon}{n_filas}")

    anchos = {
        "RFC": 16, "NOMBRE_CONSOLIDADO": 38, "TIPO_PROVEEDOR": 14, "PERSONA_FISICA": 10,
        "BUCKET": 13, "MOTIVO": 18, "PESO_ACTIVIDAD": 8, "FECHA_ULT_COMPRA": 16,
        "MONTO_MXN_2025": 15, "MONTO_USD_2025": 15, "MONTO_PEND": 14, "DECISION_PREVIA": 15,
        "FLAGS_CAMBIADOS": 20, "ESTADO_SUGERIDO": 15, "ESTADO": 14, "RAZON": 42,
        "COMENTARIO": 30, "RUN_ID_AL_DECIDIR": 26, "HASH_AL_DECIDIR": 12,
    }
    for j, nombre in enumerate(todos, start=1):
        ws.column_dimensions[get_column_letter(j)].width = anchos.get(nombre, 11)
    for nombre in COLS_CONTROL:
        ws.column_dimensions[get_column_letter(todos.index(nombre) + 1)].hidden = True

    ws.freeze_panes = ws.cell(row=2, column=len(COLS_CONTEXTO) + 1)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(todos))}{n_filas}"
    ws.protection.sheet = True
    ws.protection.autoFilter = False
    ws.protection.sort = False

    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(salida)
    return salida


def generar(dominios=("ST", "RSS"), destino: Path | None = None) -> list[str]:
    """Genera un Excel de captura por sociedad. Devuelve mensajes de resultado."""
    msgs = []
    for dominio in dominios:
        run_id, filas = fetch_datos(dominio)
        if not filas:
            msgs.append(f"{dominio}: sin proveedores por decidir.")
            continue
        ruta = construir_excel(dominio, run_id, filas, destino)
        msgs.append(f"{len(filas)} proveedores {dominio} -> {ruta.name}")
    return msgs
