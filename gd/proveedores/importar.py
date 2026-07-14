"""Lectura del Excel de captura de PROVEEDORES y UPSERT de decisiones (por RFC)."""

from __future__ import annotations

from openpyxl import load_workbook

from gd.conexion import get_connection

TABLA = "GD_DECISIONES_PROVEEDORES"
ESTADOS_VALIDOS = {"MIGRAR", "DESCARTAR", "PENDIENTE"}

HOJA = "Decisiones"
COL_RFC, COL_DOM = "RFC", "DOMINIO"
COL_ESTADO, COL_RAZON, COL_COMENT = "ESTADO", "RAZON", "COMENTARIO"
COL_RUN, COL_HASH = "RUN_ID_AL_DECIDIR", "HASH_AL_DECIDIR"


def leer_decisiones(ruta):
    wb = load_workbook(ruta, data_only=True)
    if HOJA not in wb.sheetnames:
        raise ValueError(f"El archivo no tiene la hoja '{HOJA}'. ¿Es el Excel de captura de proveedores?")
    ws = wb[HOJA]
    hdr = {}
    for idx, cell in enumerate(ws[1]):
        if cell.value is not None:
            hdr[str(cell.value).strip()] = idx
    faltan = [c for c in (COL_RFC, COL_ESTADO, COL_RAZON, COL_RUN, COL_HASH) if c not in hdr]
    if faltan:
        raise ValueError(f"Al Excel le faltan columnas requeridas: {', '.join(faltan)}")
    tiene_dom, tiene_coment = COL_DOM in hdr, COL_COMENT in hdr

    def val(row, col):
        v = row[hdr[col]].value if col in hdr else None
        if v is None:
            return None
        v = str(v).strip()
        return v or None

    decisiones, errores, sin_decidir = [], [], 0
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if all(c.value in (None, "") for c in row):
            continue
        estado = val(row, COL_ESTADO)
        if not estado:
            sin_decidir += 1
            continue
        estado = estado.upper()
        rfc = val(row, COL_RFC)
        rfc = rfc.upper() if rfc else None
        razon = val(row, COL_RAZON)
        coment = val(row, COL_COMENT) if tiene_coment else None
        run, hsh = val(row, COL_RUN), val(row, COL_HASH)
        dom = (val(row, COL_DOM) or "").upper() or None
        etq = f"Fila {i}" + (f" (RFC {rfc})" if rfc else "")
        if estado not in ESTADOS_VALIDOS:
            errores.append(f"{etq}: ESTADO inválido '{estado}'.")
            continue
        if estado != "PENDIENTE" and not razon:
            errores.append(f"{etq}: falta RAZÓN (obligatoria si ESTADO no es PENDIENTE).")
            continue
        if not (rfc and run and hsh):
            errores.append(f"{etq}: faltan datos de control (RFC/run/hash).")
            continue
        decisiones.append({"rfc": rfc, "dom": dom, "estado": estado, "razon": razon,
                           "coment": coment, "run": run, "hash": hsh})
    return decisiones, errores, sin_decidir


_MERGE = f"""
MERGE INTO {TABLA} d
USING (SELECT :rfc AS RFC FROM dual) s
   ON (d.RFC = s.RFC)
WHEN MATCHED THEN UPDATE SET
     ESTADO=:estado, RAZON=:razon, COMENTARIO=:coment, DECIDIDO_POR=:quien, ROL=:rol,
     FECHA_DECISION=SYSDATE, HASH_AL_DECIDIR=:hash, RUN_ID_AL_DECIDIR=:run, DOMINIO=:dom
WHEN NOT MATCHED THEN INSERT
     (RFC, DOMINIO, ESTADO, RAZON, COMENTARIO, DECIDIDO_POR, ROL,
      FECHA_DECISION, HASH_AL_DECIDIR, RUN_ID_AL_DECIDIR)
     VALUES (:rfc, :dom, :estado, :razon, :coment, :quien, :rol, SYSDATE, :hash, :run)
"""


def guardar_decisiones(decisiones, quien, rol=None):
    with get_connection() as con:
        cur = con.cursor()
        cur.execute(f"SELECT RFC FROM {TABLA}")
        existentes = {r[0] for r in cur.fetchall()}
        nuevas = sum(1 for x in decisiones if x["rfc"] not in existentes)
        cur.executemany(_MERGE, [dict(x, quien=(quien or "")[:60], rol=rol) for x in decisiones])
        con.commit()
    return nuevas, len(decisiones) - nuevas
