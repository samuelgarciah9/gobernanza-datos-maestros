"""Genera el reporte de avance en Excel (compartible)."""

from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from gd.conexion import BASE
from gd.datos import cargar_datos, cargar_huerfanos_sap


def generar_reporte() -> Path:
    d = cargar_datos()
    wb = Workbook()
    hf, hfill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1F4E78")

    def escribe(ws, encabezados, filas):
        for j, h in enumerate(encabezados, start=1):
            c = ws.cell(row=1, column=j, value=h)
            c.font, c.fill = hf, hfill
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[chr(64 + j)].width = max(12, len(str(h)) + 2)
        for i, fila in enumerate(filas, start=2):
            for j, v in enumerate(fila, start=1):
                ws.cell(row=i, column=j, value=v)

    ws1 = wb.active
    ws1.title = "Avance"
    escribe(ws1, ["Dominio", "Figura", "Total", "Decididos", "Pendientes", "% Avance", "Migrar", "Descartar"],
            [(r["dominio"], r["figura"], r["total"], r["decididos"], r["pendientes"],
              round(r["pct"], 1), r["migrar"], r["descartar"]) for r in d["avance"]])
    escribe(wb.create_sheet("Buckets"), ["Bucket", "Materiales"], list(d["buckets"].items()))
    escribe(wb.create_sheet("SALIO"), ["Atencion", "Materiales"],
            list(d["salio"].items()) or [("(ninguno)", 0)])

    sap = d.get("sap")
    if sap:
        escribe(wb.create_sheet("Avance SAP"),
                ["Dominio", "Figura", "Universo", "En SAP", "% Cobertura"],
                [(r["dominio"], r["figura"], r["universo"], r["en_sap"], round(r["pct"], 1))
                 for r in sap["por_dominio"]]
                + [("TOTAL", "", sap["universo"], sap["en_sap"], round(sap["pct"], 1))])
        cmp = sap["cumplimiento"]
        escribe(wb.create_sheet("Reconciliacion SAP"), ["Indicador", "Materiales"],
                [("Universo maestro", sap["universo"]),
                 ("En SAP (cobertura)", sap["en_sap"]),
                 ("Decididos MIGRAR", cmp["migrar"]),
                 ("MIGRAR ya en SAP (cumplimiento)", cmp["en_sap"]),
                 ("MIGRAR por cargar", cmp["por_cargar"]),
                 ("Cargado sin decision", sap["sin_decision"]),
                 ("Contradicciones (en SAP + DESCARTAR)", sap["contradicciones"]),
                 ("Huerfanos (en SAP fuera de universo)", sap["huerfanos"])])
        huer = cargar_huerfanos_sap()
        escribe(wb.create_sheet("Huerfanos SAP"), ["Material", "Dominio", "Descripcion"],
                [(h["material"], h["dominio"], h["descripcion"]) for h in huer]
                or [("(ninguno)", "", "")])

    marca = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    ruta = BASE / "entregables" / f"reporte_avance_{marca}.xlsx"
    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta)
    return ruta
