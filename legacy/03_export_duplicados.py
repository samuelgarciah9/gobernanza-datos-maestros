"""Exporta los duplicados ST<->RSS a un Excel controlado para decision del negocio.

Fuente : vista V_GD_MATERIALES_DUP_ST_RSS (servidor ERP-PROD).
Salida : entregables/duplicados_ST_RSS_para_decision.xlsx

El Excel trae:
  - Columnas de COMPARACION (referencia, bloqueadas): datos de cada material en
    ST vs RSS lado a lado.
  - Columnas de DECISION (editables, con listas desplegables) que llena el negocio:
      ES_MISMO_MATERIAL (SI/NO), MIGRAR_A (ST/RSS/AMBOS/DESCARTAR),
      RAZON, DECIDIDO_POR, FECHA_DECISION.

Uso:
    python 03_export_duplicados.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from connection import PROJECT_ROOT, get_connection

SALIDA = PROJECT_ROOT / "entregables" / "duplicados_ST_RSS_para_decision.xlsx"

# Columnas de decision que llena el negocio (encabezado -> ancho)
COLS_DECISION = [
    "ES_MISMO_MATERIAL",   # SI = mismo material fisico | NO = codigos distintos
    "MIGRAR_A",            # ST / RSS / AMBOS / DESCARTAR
    "RAZON",               # texto libre
    "DECIDIDO_POR",        # nombre de quien decide
    "FECHA_DECISION",      # fecha
]

# Estilos
AZUL = "1F4E78"
GRIS = "808080"
VERDE = "375623"
BORDE = Border(*(Side(style="thin", color="D9D9D9"),) * 4)


def fetch_duplicados():
    """Devuelve (encabezados, filas) de la vista de duplicados."""
    with get_connection() as con:
        with con.cursor() as cur:
            cur.execute(
                "SELECT * FROM V_GD_MATERIALES_DUP_ST_RSS "
                "ORDER BY NUMERO_PRODUCTO_ANTIGUO"
            )
            encabezados = [d[0] for d in cur.description]
            filas = cur.fetchall()
    return encabezados, filas


def construir_excel(encabezados, filas) -> Path:
    wb = Workbook()

    # -------- Hoja de instrucciones --------
    ins = wb.active
    ins.title = "Instrucciones"
    lineas = [
        ("Duplicados ST <-> RSS  -  Decision del negocio", True),
        ("", False),
        ("Estos materiales existen con el mismo codigo en AMBAS especialidades (ST y RSS).", False),
        ("La migracion a SAP se hace por separado por dominio, por eso el negocio debe decidir", False),
        ("como tratar cada uno. Las columnas grises son solo de referencia (no se editan).", False),
        ("", False),
        ("Solo llene las columnas verdes:", True),
        ("  - ES_MISMO_MATERIAL:  SI = es el mismo material fisico en ambos dominios.", False),
        ("                        NO = son materiales distintos que comparten codigo.", False),
        ("  - MIGRAR_A:  a que dominio se migra -> ST / RSS / AMBOS / DESCARTAR.", False),
        ("               (AMBOS solo si ES_MISMO_MATERIAL = NO).", False),
        ("  - RAZON:  breve justificacion de la decision.", False),
        ("  - DECIDIDO_POR:  su nombre.", False),
        ("  - FECHA_DECISION:  fecha en que decide.", False),
        ("", False),
        ("Guia rapida: la columna DOMINIO_MAS_ACTIVO sugiere donde hay mas actividad,", False),
        ("pero NO sustituye su criterio.", False),
    ]
    for i, (texto, negrita) in enumerate(lineas, start=1):
        c = ins.cell(row=i, column=1, value=texto)
        c.font = Font(bold=negrita, size=12 if (negrita and i == 1) else 11,
                      color=AZUL if i == 1 else "000000")
    ins.column_dimensions["A"].width = 95

    # -------- Hoja de datos --------
    ws = wb.create_sheet("Duplicados")
    todos = encabezados + COLS_DECISION
    n_ref = len(encabezados)

    # Encabezados
    for j, nombre in enumerate(todos, start=1):
        c = ws.cell(row=1, column=j, value=nombre)
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.fill = PatternFill("solid", fgColor=GRIS if j <= n_ref else VERDE)
        c.border = BORDE

    # Datos de referencia
    for i, fila in enumerate(filas, start=2):
        for j, valor in enumerate(fila, start=1):
            c = ws.cell(row=i, column=j, value=valor)
            c.border = BORDE
            c.protection = Protection(locked=True)  # referencia: bloqueada
        # celdas de decision: vacias y desbloqueadas
        for j in range(n_ref + 1, len(todos) + 1):
            c = ws.cell(row=i, column=j)
            c.border = BORDE
            c.protection = Protection(locked=False)

    n_filas = len(filas) + 1  # incluye encabezado

    # Listas desplegables
    def add_dropdown(col_nombre, opciones):
        idx = todos.index(col_nombre) + 1
        letra = get_column_letter(idx)
        dv = DataValidation(type="list", formula1=f'"{opciones}"', allow_blank=True)
        dv.error = "Elija un valor de la lista."
        dv.prompt = f"Opciones: {opciones}"
        ws.add_data_validation(dv)
        dv.add(f"{letra}2:{letra}{n_filas}")

    add_dropdown("ES_MISMO_MATERIAL", "SI,NO")
    add_dropdown("MIGRAR_A", "ST,RSS,AMBOS,DESCARTAR")

    # Anchos de columna
    for j, nombre in enumerate(todos, start=1):
        letra = get_column_letter(j)
        if nombre.startswith("DESCRIPCION"):
            ws.column_dimensions[letra].width = 32
        elif nombre in ("RAZON",):
            ws.column_dimensions[letra].width = 30
        elif nombre == "NUMERO_PRODUCTO_ANTIGUO":
            ws.column_dimensions[letra].width = 22
        else:
            ws.column_dimensions[letra].width = 16

    # Formato fecha en FECHA_DECISION
    fidx = todos.index("FECHA_DECISION") + 1
    fletra = get_column_letter(fidx)
    for i in range(2, n_filas + 1):
        ws[f"{fletra}{i}"].number_format = "DD/MM/YYYY"

    # Congelar encabezado + filtro + proteccion (sin password: solo evita edicion accidental)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(todos))}{n_filas}"
    ws.protection.sheet = True
    ws.protection.autoFilter = False  # permitir filtrar
    ws.protection.sort = False        # permitir ordenar

    SALIDA.parent.mkdir(parents=True, exist_ok=True)
    wb.save(SALIDA)
    return SALIDA


if __name__ == "__main__":
    encabezados, filas = fetch_duplicados()
    ruta = construir_excel(encabezados, filas)
    print(f"OK - {len(filas)} duplicados exportados a:")
    print(f"     {ruta}")
