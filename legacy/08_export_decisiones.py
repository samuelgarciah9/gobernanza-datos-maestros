"""Paso 2a - EXPORTA el snapshot de un dominio a un Excel de captura de decisiones.

Fuente : GD_SNAPSHOT_MATERIALES (ultimo RUN_ID del dominio) + GD_CAT_RAZONES.
Salida : entregables/decisiones_<DOM>_<RUN_ID>_para_captura.xlsx

El Excel trae tres bloques de columnas:
  - CONTEXTO (gris, bloqueado): datos y banderas del material para decidir informado.
  - SUGERENCIA (amarillo): ESTADO_SUGERIDO calculado por reglas del snapshot.
  - DECISION (verde, editable): ESTADO (desplegable), RAZON (desplegable del
    catalogo) y COMENTARIO. Es lo unico que llena la persona.
  - CONTROL (oculto): RUN_ID_AL_DECIDIR y HASH_AL_DECIDIR amarran cada decision
    a la foto exacta que se vio (para el merge del Paso 3).

Uso:
    python 08_export_decisiones.py            # dominio ST (piloto)
    python 08_export_decisiones.py RSS        # otro dominio
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from connection import PROJECT_ROOT, get_connection

VISTA_MERGE = "V_GD_MERGE_MATERIALES"   # fuente: buckets del Paso 3 sobre el ultimo snapshot
TABLA_CAT = "GD_CAT_RAZONES"
ESTADOS = ("MIGRAR", "DESCARTAR", "PENDIENTE")
# Solo se exportan los buckets que necesitan atencion del revisor.
BUCKETS_CAPTURA = ("POR_DECIDIR", "RE_REVISAR")

# Columnas de CONTEXTO que se muestran (encabezado en el snapshot -> etiqueta visible).
# Las banderas HAS_* se traducen a Si/No al escribir.
# TIPO_MATERIAL es DERIVADA (no existe en el snapshot): se calcula de PT_PROD_LINE.
COLS_CONTEXTO = [
    "NUMERO_PRODUCTO_ANTIGUO",
    "TIPO_MATERIAL",         # PRODUCTIVO / NO_PRODUCTIVO (derivada de PT_PROD_LINE)
    "BUCKET",                # POR_DECIDIR / RE_REVISAR (Paso 3)
    "MOTIVO",                # POSIBLE_MIGRAR / POSIBLE_DESCARTAR (solo en RE_REVISAR)
    "PT_PROD_LINE",
    "DESCRIPCION",
    "GRUPO_PRODUCTO",
    "PT_GROUP",
    "PT_PART_TYPE",
    "EXIST_TOTAL",
    "ULTIMA_TRANSACCION",
    "ULTIMO_TIPO_TXN",
    "DIAS_SIN_MOVIMIENTO",   # antiguedad real (NULL/vacio = nunca tuvo transaccion)
    # --- banderas de actividad (cada una vale 1, sin pesos) ---
    "HAS_EXIST",
    "HAS_PO",
    "HAS_SO",
    "HAS_WO",
    "HAS_INV_SEG",
    "HAS_DIST",
    "HAS_RET",
    "PESO_ACTIVIDAD",        # numero de senales activas (0..8, suma plana sin pesos)
    # --- contexto de reconciliacion (se llenan solo en RE_REVISAR) ---
    "DECISION_PREVIA",       # que se habia decidido antes
    "FLAGS_CAMBIADOS",       # que banderas cambiaron (+aparecio / -desaparecio)
]
COLS_BANDERA = {
    "HAS_EXIST", "HAS_PO", "HAS_SO", "HAS_WO", "HAS_INV_SEG", "HAS_DIST", "HAS_RET",
}
# Columnas DERIVADAS (no se piden al SELECT; se calculan en Python).
DERIVADAS = {"TIPO_MATERIAL"}

# Segmentacion del piloto: dos figuras revisan por separado.
#   NO_PRODUCTIVO = PT_PROD_LINE 'REF' ; PRODUCTIVO = el resto.
SEGMENTOS = ("PRODUCTIVO", "NO_PRODUCTIVO")


def tipo_material(fila: dict) -> str:
    pl = (fila.get("PT_PROD_LINE") or "").strip().upper()
    return "NO_PRODUCTIVO" if pl == "REF" else "PRODUCTIVO"

# Señales FIRMES de negocio (stock / orden / inv. seguridad). Se usan SOLO para la
# heurística de sugerencia; NO son pesos: PESO_ACTIVIDAD sigue siendo suma plana.
SENALES_FIRMES = ("HAS_EXIST", "HAS_PO", "HAS_SO", "HAS_WO", "HAS_INV_SEG")

COL_SUGERENCIA = "ESTADO_SUGERIDO"
COLS_DECISION = ["ESTADO", "RAZON", "COMENTARIO"]
# Control (ocultas): PT_DOMAIN completa la llave del upsert (material, dominio).
COLS_CONTROL = ["PT_DOMAIN", "RUN_ID_AL_DECIDIR", "HASH_AL_DECIDIR"]

# Estilos (reusa la paleta de 03_export_duplicados.py)
AZUL = "1F4E78"
GRIS = "808080"
AMARILLO = "BF8F00"
VERDE = "375623"
BORDE = Border(*(Side(style="thin", color="D9D9D9"),) * 4)


def sugerir_estado(fila: dict) -> str:
    """Sugerencia por TIPO de señal (heurística de negocio). NO sustituye al humano.

      - Cualquier señal FIRME (stock/PO/SO/WO/inv.seg) -> MIGRAR.
      - Solo señal ligera (DIST/RET) o txn reciente     -> PENDIENTE (revisar).
      - Sin ninguna señal                               -> DESCARTAR (no ocurre en la vista).
    """
    if any(fila.get(f) for f in SENALES_FIRMES):
        return "MIGRAR"
    if fila.get("HAS_DIST") or fila.get("HAS_RET") or fila.get("HAS_ULT_TXN"):
        return "PENDIENTE"
    return "DESCARTAR"


def fetch_datos(dominio: str):
    """Devuelve (run_id, razones, filas[dict]) de la vista de merge para el dominio.

    Solo trae los buckets que necesitan atencion (POR_DECIDIR y RE_REVISAR);
    los VIGENTE y SALIO quedan fuera del Excel de captura.
    """
    with get_connection() as con:
        cur = con.cursor()
        cur.execute(
            f"SELECT RAZON FROM {TABLA_CAT} WHERE ACTIVO='S' ORDER BY ESTADO_SUGERIDO, RAZON"
        )
        razones = [r[0] for r in cur.fetchall()]

        # Columnas necesarias (sin derivadas). Se leen de la vista de merge.
        reales = [c for c in COLS_CONTEXTO if c not in DERIVADAS]
        necesarias = list(dict.fromkeys(
            reales + ["PT_PROD_LINE", "HAS_ULT_TXN", "BANDERAS_HASH", "PT_DOMAIN", "RUN_ID"]
        ))
        marcadores = ", ".join(f"'{b}'" for b in BUCKETS_CAPTURA)
        cur.execute(
            f"SELECT {', '.join(necesarias)} FROM {VISTA_MERGE} "
            f"WHERE PT_DOMAIN = :d AND BUCKET IN ({marcadores}) "
            # RE_REVISAR primero (mas urgente), luego por material
            f"ORDER BY CASE WHEN BUCKET = 'RE_REVISAR' THEN 0 ELSE 1 END, "
            f"NUMERO_PRODUCTO_ANTIGUO",
            d=dominio,
        )
        cols = [c[0] for c in cur.description]
        filas = [dict(zip(cols, vals)) for vals in cur.fetchall()]

    # Deriva TIPO_MATERIAL (segmento del piloto) por fila.
    for f in filas:
        f["TIPO_MATERIAL"] = tipo_material(f)
    run_id = filas[0]["RUN_ID"] if filas else None
    return run_id, razones, filas


def construir_excel(dominio: str, tipo: str, run_id: str, razones, filas) -> Path:
    salida = (PROJECT_ROOT / "entregables"
              / f"decisiones_{dominio}_{tipo}_{run_id}_para_captura.xlsx")
    wb = Workbook()

    # -------- Hoja de instrucciones --------
    ins = wb.active
    ins.title = "Instrucciones"
    lineas = [
        (f"Captura de decisiones de migracion  -  Dominio {dominio}  -  {tipo}", True),
        ("", False),
        (f"Foto (snapshot): {run_id}   |   Materiales: {len(filas)}", False),
        ("", False),
        ("Cada fila es un material. Decida si se MIGRA o se DESCARTA (o queda PENDIENTE).", False),
        ("Las columnas grises son referencia (no se editan).", False),
        ("", False),
        ("Solo llene las columnas verdes:", True),
        ("  - ESTADO:  elija de la lista -> MIGRAR / DESCARTAR / PENDIENTE.", False),
        ("  - RAZON:  elija una razon del catalogo (obligatoria si ESTADO no es PENDIENTE).", False),
        ("  - COMENTARIO:  nota libre opcional.", False),
        ("", False),
        ("La columna amarilla ESTADO_SUGERIDO es una recomendacion automatica basada", False),
        ("en las banderas del material (existencias, ordenes, obsoleto). Puede confirmarla", False),
        ("o cambiarla; NO sustituye su criterio.", False),
        ("", False),
        ("BUCKET indica el estado del material en el proceso:", True),
        ("  - POR_DECIDIR: material nuevo o sin decision -> decida normal.", False),
        ("  - RE_REVISAR: ya tenia decision y ALGO cambio. Vea MOTIVO (posible migrar/descartar)", False),
        ("    y FLAGS_CAMBIADOS (que se movio), y reconsidere la decision previa.", False),
        ("", False),
        ("NO borre ni edite las dos ultimas columnas de control (van ocultas).", False),
    ]
    for i, (texto, negrita) in enumerate(lineas, start=1):
        c = ins.cell(row=i, column=1, value=texto)
        c.font = Font(bold=negrita, size=12 if (negrita and i == 1) else 11,
                      color=AZUL if i == 1 else "000000")
    ins.column_dimensions["A"].width = 95

    # -------- Hoja auxiliar con el catalogo de razones (para el desplegable) --------
    cat = wb.create_sheet("Catalogo")
    cat["A1"] = "RAZONES (no editar)"
    cat["A1"].font = Font(bold=True)
    for i, r in enumerate(razones, start=2):
        cat[f"A{i}"] = r
    cat.column_dimensions["A"].width = 60
    cat.sheet_state = "hidden"

    # -------- Hoja de datos --------
    ws = wb.create_sheet("Decisiones")
    todos = COLS_CONTEXTO + [COL_SUGERENCIA] + COLS_DECISION + COLS_CONTROL
    n_ctx = len(COLS_CONTEXTO)
    idx_sug = n_ctx  # 0-based indice de la columna de sugerencia

    # Encabezados con color por bloque
    for j, nombre in enumerate(todos, start=1):
        c = ws.cell(row=1, column=j, value=nombre)
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if nombre == COL_SUGERENCIA:
            fill = AMARILLO
        elif nombre in COLS_DECISION:
            fill = VERDE
        else:  # contexto y control
            fill = GRIS
        c.fill = PatternFill("solid", fgColor=fill)
        c.border = BORDE

    # Datos
    for i, fila in enumerate(filas, start=2):
        sugerido = sugerir_estado(fila)
        for j, nombre in enumerate(todos, start=1):
            if nombre in COLS_DECISION:
                valor = None  # lo llena la persona
            elif nombre == COL_SUGERENCIA:
                valor = sugerido
            elif nombre == "RUN_ID_AL_DECIDIR":
                valor = run_id
            elif nombre == "HASH_AL_DECIDIR":
                valor = fila.get("BANDERAS_HASH")
            elif nombre in COLS_BANDERA:
                valor = "Si" if fila.get(nombre) else "No"
            else:
                valor = fila.get(nombre)
            c = ws.cell(row=i, column=j, value=valor)
            c.border = BORDE
            # Editable solo el bloque de decision
            c.protection = Protection(locked=nombre not in COLS_DECISION)

    n_filas = len(filas) + 1  # incluye encabezado

    # Desplegables
    letra_estado = get_column_letter(todos.index("ESTADO") + 1)
    dv_estado = DataValidation(type="list", formula1=f'"{",".join(ESTADOS)}"', allow_blank=True)
    dv_estado.error = "Elija un estado de la lista."
    dv_estado.prompt = f"Opciones: {', '.join(ESTADOS)}"
    ws.add_data_validation(dv_estado)
    dv_estado.add(f"{letra_estado}2:{letra_estado}{n_filas}")

    letra_razon = get_column_letter(todos.index("RAZON") + 1)
    ult_cat = len(razones) + 1
    dv_razon = DataValidation(
        type="list", formula1=f"Catalogo!$A$2:$A${ult_cat}", allow_blank=True
    )
    dv_razon.error = "Elija una razon del catalogo."
    ws.add_data_validation(dv_razon)
    dv_razon.add(f"{letra_razon}2:{letra_razon}{n_filas}")

    # Anchos
    anchos = {
        "NUMERO_PRODUCTO_ANTIGUO": 22, "TIPO_MATERIAL": 15, "PT_PROD_LINE": 13,
        "BUCKET": 13, "MOTIVO": 18, "DECISION_PREVIA": 15, "FLAGS_CAMBIADOS": 22,
        "DESCRIPCION": 34, "GRUPO_PRODUCTO": 20,
        "PT_GROUP": 16, "PT_PART_TYPE": 16, "ULTIMA_TRANSACCION": 16,
        "ULTIMO_TIPO_TXN": 16, "DIAS_SIN_MOVIMIENTO": 14, "PESO_ACTIVIDAD": 14,
        "ESTADO_SUGERIDO": 15, "ESTADO": 14,
        "RAZON": 42, "COMENTARIO": 30, "RUN_ID_AL_DECIDIR": 28, "HASH_AL_DECIDIR": 14,
    }
    for j, nombre in enumerate(todos, start=1):
        ws.column_dimensions[get_column_letter(j)].width = anchos.get(nombre, 11)

    # Ocultar columnas de control
    for nombre in COLS_CONTROL:
        ws.column_dimensions[get_column_letter(todos.index(nombre) + 1)].hidden = True

    # Congelar, filtro y proteccion (sin password: evita edicion accidental)
    ws.freeze_panes = ws.cell(row=2, column=n_ctx + 1)  # fija contexto a la izquierda
    ws.auto_filter.ref = f"A1:{get_column_letter(len(todos))}{n_filas}"
    ws.protection.sheet = True
    ws.protection.autoFilter = False
    ws.protection.sort = False

    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(salida)
    return salida


if __name__ == "__main__":
    # Uso: python 08_export_decisiones.py [ST|RSS] [PRODUCTIVO|NO_PRODUCTIVO]
    dominio = (sys.argv[1] if len(sys.argv) > 1 else "ST").upper()
    if dominio not in ("ST", "RSS"):
        sys.exit(f"Dominio no valido: {dominio}. Use ST o RSS.")
    tipo_arg = sys.argv[2].upper() if len(sys.argv) > 2 else None
    if tipo_arg and tipo_arg not in SEGMENTOS:
        sys.exit(f"Tipo no valido: {tipo_arg}. Use {SEGMENTOS}.")

    run_id, razones, filas = fetch_datos(dominio)
    objetivo = (tipo_arg,) if tipo_arg else SEGMENTOS

    # Un Excel por figura (segmento). Se omite el segmento sin materiales.
    for seg in objetivo:
        sub = [f for f in filas if f["TIPO_MATERIAL"] == seg]
        if not sub:
            print(f"  ({dominio} / {seg}: 0 materiales, se omite)")
            continue
        ruta = construir_excel(dominio, seg, run_id, razones, sub)
        print(f"OK - {len(sub)} materiales {dominio} / {seg} ({run_id}) -> {ruta}")
