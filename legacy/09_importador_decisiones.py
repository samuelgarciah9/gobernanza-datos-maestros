"""Importador de Decisiones  -  App de escritorio para los revisores.

Objetivo: que cada figura (productivo / no productivo) cargue SU Excel de captura
a la tabla GD_DECISIONES_MATERIALES de forma AUTONOMA, sin intervencion tecnica.

Como funciona:
  1. El revisor escribe su nombre y elige su rol (se recuerda para la proxima vez).
  2. Selecciona su archivo Excel lleno (el generado por 08_export_decisiones.py).
  3. Da clic en "Cargar decisiones": la app valida y hace UPSERT (MERGE) por
     (material, dominio). Es re-ejecutable: pueden cargar cuantas veces quieran,
     conforme van decidiendo; si cambian una decision, se actualiza.

Lanzar:
    doble clic en "Importador de decisiones.bat"
    o:  pythonw 09_importador_decisiones.py

Requisitos en la maquina del revisor: Python + oracledb + openpyxl + python-dotenv,
el Oracle Instant Client (modo Thick) y el archivo .env con las credenciales.
"""

from __future__ import annotations

import json
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, ttk

from openpyxl import load_workbook

from connection import PROJECT_ROOT, get_connection

TABLA = "GD_DECISIONES_MATERIALES"
ESTADOS_VALIDOS = {"MIGRAR", "DESCARTAR", "PENDIENTE"}
ROLES = ["COMPRAS", "INVENTARIOS", "INGENIERIA"]
CONFIG = PROJECT_ROOT / ".importador_config.json"
CARPETA_ENTREGABLES = PROJECT_ROOT / "entregables"

# Encabezados esperados en la hoja "Decisiones"
HOJA = "Decisiones"
COL_MAT, COL_DOM = "NUMERO_PRODUCTO_ANTIGUO", "PT_DOMAIN"
COL_ESTADO, COL_RAZON, COL_COMENT = "ESTADO", "RAZON", "COMENTARIO"
COL_RUN, COL_HASH = "RUN_ID_AL_DECIDIR", "HASH_AL_DECIDIR"


# ----------------------------------------------------------------------------
# Logica de negocio (independiente de la interfaz)
# ----------------------------------------------------------------------------
def leer_decisiones(ruta: str):
    """Lee el Excel y devuelve (decisiones, errores, sin_decidir).

    Solo toma filas con ESTADO lleno. Valida estado/razon/llaves de control.
    """
    wb = load_workbook(ruta, data_only=True)
    if HOJA not in wb.sheetnames:
        raise ValueError(
            f"El archivo no tiene la hoja '{HOJA}'.\n"
            "Verifique que sea el Excel de captura de decisiones."
        )
    ws = wb[HOJA]

    hdr = {}
    for idx, cell in enumerate(ws[1]):
        if cell.value is not None:
            hdr[str(cell.value).strip()] = idx

    faltan = [c for c in (COL_MAT, COL_ESTADO, COL_RAZON, COL_RUN, COL_HASH) if c not in hdr]
    if faltan:
        raise ValueError(f"Al Excel le faltan columnas requeridas: {', '.join(faltan)}")

    tiene_dom = COL_DOM in hdr
    tiene_coment = COL_COMENT in hdr

    def val(row, col):
        v = row[hdr[col]].value if col in hdr else None
        if v is None:
            return None
        v = str(v).strip()
        return v or None

    decisiones, errores, sin_decidir = [], [], 0
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if all(c.value in (None, "") for c in row):
            continue  # fila totalmente vacia

        estado = val(row, COL_ESTADO)
        if not estado:
            sin_decidir += 1
            continue
        estado = estado.upper()

        mat = val(row, COL_MAT)
        razon = val(row, COL_RAZON)
        coment = val(row, COL_COMENT) if tiene_coment else None
        run = val(row, COL_RUN)
        hsh = val(row, COL_HASH)

        dom = val(row, COL_DOM) if tiene_dom else None
        if not dom and run:  # fallback: RUN_ST_... / RUN_RSS_...
            partes = run.split("_")
            dom = partes[1].upper() if len(partes) >= 2 else None
        dom = dom.upper() if dom else None

        etq = f"Fila {i}" + (f" (material {mat})" if mat else "")
        if estado not in ESTADOS_VALIDOS:
            errores.append(f"{etq}: ESTADO invalido '{estado}'.")
            continue
        if estado != "PENDIENTE" and not razon:
            errores.append(f"{etq}: falta RAZON (obligatoria si ESTADO no es PENDIENTE).")
            continue
        if not (mat and dom and run and hsh):
            errores.append(f"{etq}: faltan datos de control (material/dominio/run/hash).")
            continue

        decisiones.append({
            "mat": mat, "dom": dom, "estado": estado, "razon": razon,
            "coment": coment, "run": run, "hash": hsh,
        })

    return decisiones, errores, sin_decidir


MERGE_SQL = f"""
MERGE INTO {TABLA} d
USING (SELECT :mat AS NUMERO_PRODUCTO_ANTIGUO, :dom AS PT_DOMAIN FROM dual) s
   ON (d.NUMERO_PRODUCTO_ANTIGUO = s.NUMERO_PRODUCTO_ANTIGUO AND d.PT_DOMAIN = s.PT_DOMAIN)
WHEN MATCHED THEN UPDATE SET
     ESTADO = :estado, RAZON = :razon, COMENTARIO = :coment,
     DECIDIDO_POR = :quien, ROL = :rol, FECHA_DECISION = SYSDATE,
     HASH_AL_DECIDIR = :hash, RUN_ID_AL_DECIDIR = :run
WHEN NOT MATCHED THEN INSERT
     (NUMERO_PRODUCTO_ANTIGUO, PT_DOMAIN, ESTADO, RAZON, COMENTARIO,
      DECIDIDO_POR, ROL, FECHA_DECISION, HASH_AL_DECIDIR, RUN_ID_AL_DECIDIR)
     VALUES (:mat, :dom, :estado, :razon, :coment, :quien, :rol, SYSDATE, :hash, :run)
"""


def guardar_decisiones(decisiones, quien, rol):
    """UPSERT en la tabla. Devuelve (nuevas, actualizadas)."""
    with get_connection() as con:
        cur = con.cursor()
        # Clasifica nuevas vs actualizadas leyendo las llaves existentes (tabla chica)
        cur.execute(f"SELECT NUMERO_PRODUCTO_ANTIGUO, PT_DOMAIN FROM {TABLA}")
        existentes = {(m, d) for m, d in cur.fetchall()}
        nuevas = sum(1 for x in decisiones if (x["mat"], x["dom"]) not in existentes)
        actualizadas = len(decisiones) - nuevas

        binds = [dict(x, quien=quien[:60], rol=rol) for x in decisiones]
        cur.executemany(MERGE_SQL, binds)
        con.commit()
    return nuevas, actualizadas


def cargar_config():
    try:
        return json.loads(CONFIG.read_text(encoding="utf-8"))
    except Exception:
        return {}


def guardar_config(nombre, rol):
    try:
        CONFIG.write_text(json.dumps({"nombre": nombre, "rol": rol}), encoding="utf-8")
    except Exception:
        pass


# ----------------------------------------------------------------------------
# Interfaz grafica
# ----------------------------------------------------------------------------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Importador de Decisiones - Gobernanza de Datos")
        self.geometry("720x560")
        self.minsize(640, 480)
        self.archivo = None

        cfg = cargar_config()
        pad = {"padx": 10, "pady": 6}

        tk.Label(self, text="Cargar decisiones a la base de datos",
                 font=("Segoe UI", 14, "bold")).pack(anchor="w", **pad)

        # -- Revisor --
        rev = tk.LabelFrame(self, text=" Revisor ", font=("Segoe UI", 10, "bold"))
        rev.pack(fill="x", **pad)
        tk.Label(rev, text="Nombre:").grid(row=0, column=0, sticky="e", padx=6, pady=6)
        self.nombre = tk.Entry(rev, width=40)
        self.nombre.insert(0, cfg.get("nombre", ""))
        self.nombre.grid(row=0, column=1, sticky="w", padx=6, pady=6)
        tk.Label(rev, text="Rol:").grid(row=0, column=2, sticky="e", padx=6, pady=6)
        self.rol = ttk.Combobox(rev, values=ROLES, state="readonly", width=16)
        self.rol.set(cfg.get("rol", ROLES[0]))
        self.rol.grid(row=0, column=3, sticky="w", padx=6, pady=6)

        # -- Archivo --
        arch = tk.LabelFrame(self, text=" Archivo de decisiones (Excel) ",
                             font=("Segoe UI", 10, "bold"))
        arch.pack(fill="x", **pad)
        self.ruta_var = tk.StringVar(value="(ningun archivo seleccionado)")
        tk.Entry(arch, textvariable=self.ruta_var, state="readonly").pack(
            side="left", fill="x", expand=True, padx=6, pady=8)
        tk.Button(arch, text="Examinar...", command=self.examinar).pack(
            side="right", padx=6, pady=8)

        # -- Boton cargar --
        self.btn = tk.Button(self, text="  Cargar decisiones  ",
                             font=("Segoe UI", 11, "bold"),
                             bg="#1F6F43", fg="white", command=self.cargar)
        self.btn.pack(**pad)

        # -- Log --
        tk.Label(self, text="Resultado:").pack(anchor="w", padx=10)
        self.log = scrolledtext.ScrolledText(self, height=14, state="disabled",
                                             font=("Consolas", 9))
        self.log.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self._log("Listo. Selecciona tu archivo y da clic en 'Cargar decisiones'.")

    def _log(self, msg):
        self.log.config(state="normal")
        self.log.insert("end", msg + "\n")
        self.log.see("end")
        self.log.config(state="disabled")
        self.update_idletasks()

    def examinar(self):
        inicial = CARPETA_ENTREGABLES if CARPETA_ENTREGABLES.exists() else PROJECT_ROOT
        ruta = filedialog.askopenfilename(
            title="Selecciona tu Excel de decisiones",
            initialdir=str(inicial),
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
        )
        if ruta:
            self.archivo = ruta
            self.ruta_var.set(ruta)
            self._log(f"\nArchivo seleccionado:\n  {Path(ruta).name}")

    def cargar(self):
        nombre = self.nombre.get().strip()
        rol = self.rol.get().strip()
        if not nombre:
            messagebox.showwarning("Falta el nombre", "Escribe tu nombre antes de cargar.")
            return
        if rol not in ROLES:
            messagebox.showwarning("Falta el rol", "Elige tu rol de la lista.")
            return
        if not self.archivo:
            messagebox.showwarning("Falta el archivo", "Selecciona tu Excel de decisiones.")
            return

        self.btn.config(state="disabled")
        try:
            self._log("\nLeyendo el archivo...")
            decisiones, errores, sin_decidir = leer_decisiones(self.archivo)

            self._log(f"  Decisiones tomadas   : {len(decisiones)}")
            self._log(f"  Sin decidir (vacias) : {sin_decidir}")
            if errores:
                self._log(f"  Filas con problema   : {len(errores)}")
                for e in errores[:25]:
                    self._log(f"     - {e}")
                if len(errores) > 25:
                    self._log(f"     ... y {len(errores) - 25} mas.")

            if not decisiones:
                self._log("\nNo hay decisiones validas que cargar.")
                messagebox.showinfo("Sin cambios", "No se cargo ninguna decision.")
                return

            self._log("\nGuardando en la base de datos...")
            nuevas, actualizadas = guardar_decisiones(decisiones, nombre, rol)
            guardar_config(nombre, rol)

            resumen = (f"{nuevas} nuevas, {actualizadas} actualizadas "
                       f"({len(decisiones)} en total).")
            self._log("LISTO: " + resumen)
            if errores:
                self._log("Nota: las filas con problema NO se cargaron; corrigelas y vuelve a cargar.")
            messagebox.showinfo("Decisiones cargadas", "Se guardaron correctamente:\n\n" + resumen)

        except Exception as e:  # noqa: BLE001
            self._log(f"\nERROR: {e}")
            messagebox.showerror(
                "Error al cargar",
                f"No se pudo completar la carga:\n\n{e}\n\n"
                "Si el problema persiste, avisa al equipo de datos.",
            )
        finally:
            self.btn.config(state="normal")


if __name__ == "__main__":
    App().mainloop()
