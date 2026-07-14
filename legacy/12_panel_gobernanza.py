"""Panel de Gobernanza  -  Herramienta del EQUIPO DE DATOS MAESTROS.

Envuelve las operaciones del proceso para que no haya que usar la linea de comandos:
  - Correr una nueva corrida (foto / snapshot)         -> 05_snapshot.py
  - Generar los Excels de captura por figura           -> 08_export_decisiones.py
  - Monitorear el avance y los buckets (en vivo)       -> vistas V_GD_MERGE_*
  - Exportar un reporte de avance compartible (Excel)

Lanzar:
    doble clic en "Panel de Gobernanza.bat"
    o:  pythonw 12_panel_gobernanza.py

Requisitos: Python + oracledb + openpyxl + python-dotenv, Instant Client y .env.
"""

from __future__ import annotations

import datetime as dt
import subprocess
import sys
import threading
import tkinter as tk
from pathlib import Path
from tkinter import messagebox, scrolledtext, ttk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from connection import PROJECT_ROOT, get_connection

PY = sys.executable
SCRIPT_SNAPSHOT = "05_snapshot.py"
SCRIPT_EXPORT = "08_export_decisiones.py"
CARPETA_ENTREGABLES = PROJECT_ROOT / "entregables"

# Paleta para el reporte Excel (se conserva)
AZUL = "1F4E78"

# --- Paleta de la interfaz (plana, moderna) ---
BG = "#EEF1F5"        # fondo ventana
HEADER = "#16324F"    # barra superior (acero)
HEADER_SUB = "#9FB3C8"
CARD = "#FFFFFF"
BORDER = "#E1E6EC"
TEXT = "#1F2937"
MUTED = "#6B7280"
GREEN = "#2E7D46"
GREEN_H = "#256B3B"
BLUE = "#1F4E78"
BLUE_H = "#17395A"
SOFT = "#E8EEF5"
SOFT_H = "#DCE6F0"
STRIPE = "#F4F7FA"


# ----------------------------------------------------------------------------
# Consultas de monitoreo (reutilizadas por el panel y por el reporte)
# ----------------------------------------------------------------------------
def fetch_buckets(cur):
    cur.execute(
        """SELECT PT_DOMAIN, BUCKET, COUNT(*)
             FROM V_GD_MERGE_MATERIALES
            GROUP BY PT_DOMAIN, BUCKET
            ORDER BY PT_DOMAIN, BUCKET"""
    )
    return cur.fetchall()


def fetch_avance(cur):
    """Por dominio y figura: total, decididos, pendientes, desglose por estado."""
    cur.execute(
        """SELECT PT_DOMAIN,
                  CASE WHEN UPPER(PT_PROD_LINE)='REF' THEN 'NO_PRODUCTIVO' ELSE 'PRODUCTIVO' END AS FIGURA,
                  COUNT(*) AS TOTAL,
                  SUM(CASE WHEN DECISION_PREVIA IN ('MIGRAR','DESCARTAR','ENRIQUECER') THEN 1 ELSE 0 END) AS DECIDIDOS,
                  SUM(CASE WHEN DECISION_PREVIA='MIGRAR'     THEN 1 ELSE 0 END) AS MIGRAR,
                  SUM(CASE WHEN DECISION_PREVIA='DESCARTAR'  THEN 1 ELSE 0 END) AS DESCARTAR,
                  SUM(CASE WHEN DECISION_PREVIA='ENRIQUECER' THEN 1 ELSE 0 END) AS ENRIQUECER
             FROM V_GD_MERGE_MATERIALES
            GROUP BY PT_DOMAIN,
                     CASE WHEN UPPER(PT_PROD_LINE)='REF' THEN 'NO_PRODUCTIVO' ELSE 'PRODUCTIVO' END
            ORDER BY PT_DOMAIN, FIGURA"""
    )
    filas = []
    for dom, fig, total, dec, mig, des, enr in cur.fetchall():
        pend = total - dec
        pct = (100.0 * dec / total) if total else 0.0
        filas.append((dom, fig, total, dec, pend, pct, mig, des, enr))
    return filas


def fetch_salio(cur):
    cur.execute(
        """SELECT ATENCION, COUNT(*) FROM V_GD_MERGE_SALIO
            GROUP BY ATENCION ORDER BY ATENCION"""
    )
    return cur.fetchall()


def generar_reporte(ruta: Path):
    """Genera un Excel de avance compartible (3 hojas)."""
    with get_connection() as con:
        cur = con.cursor()
        avance = fetch_avance(cur)
        buckets = fetch_buckets(cur)
        salio = fetch_salio(cur)

    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor=AZUL)

    def escribe(ws, encabezados, filas):
        for j, h in enumerate(encabezados, start=1):
            c = ws.cell(row=1, column=j, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")
        for i, fila in enumerate(filas, start=2):
            for j, v in enumerate(fila, start=1):
                ws.cell(row=i, column=j, value=v)
        for j, h in enumerate(encabezados, start=1):
            ws.column_dimensions[chr(64 + j)].width = max(12, len(str(h)) + 2)

    ws1 = wb.active
    ws1.title = "Avance"
    escribe(ws1, ["Dominio", "Figura", "Total", "Decididos", "Pendientes",
                  "% Avance", "Migrar", "Descartar", "Enriquecer"],
            [(d, f, t, de, p, round(pc, 1), mi, ds, en)
             for d, f, t, de, p, pc, mi, ds, en in avance])

    ws2 = wb.create_sheet("Buckets")
    escribe(ws2, ["Dominio", "Bucket", "Materiales"], buckets)

    ws3 = wb.create_sheet("SALIO")
    escribe(ws3, ["Atencion", "Materiales"], salio if salio else [("(ninguno)", 0)])

    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta)
    return ruta


# ----------------------------------------------------------------------------
# Interfaz
# ----------------------------------------------------------------------------
class Panel(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Panel de Gobernanza - Datos Maestros")
        self.geometry("960x720")
        self.minsize(880, 640)
        self.configure(bg=BG)
        self.ocupado = False

        self._init_estilos()
        self._construir_header()

        cont = tk.Frame(self, bg=BG)
        cont.pack(fill="both", expand=True, padx=18, pady=14)

        self._construir_operacion(cont)
        self._construir_monitoreo(cont)
        self._construir_registro(cont)

        self._log("Listo. 'Correr nueva corrida' toma una foto nueva; luego 'Generar Excels'.")
        self.after(300, self.actualizar)

    # -- estilos ttk --
    def _init_estilos(self):
        st = ttk.Style(self)
        try:
            st.theme_use("clam")
        except tk.TclError:
            pass
        st.configure("Treeview",
                     background=CARD, fieldbackground=CARD, foreground=TEXT,
                     rowheight=30, borderwidth=0, font=("Segoe UI", 9))
        st.configure("Treeview.Heading",
                     background=HEADER, foreground="white", relief="flat",
                     font=("Segoe UI", 9, "bold"), padding=6)
        st.map("Treeview.Heading", background=[("active", BLUE)])
        st.map("Treeview",
               background=[("selected", GREEN)], foreground=[("selected", "white")])
        st.configure("Avance.Horizontal.TProgressbar",
                     troughcolor=SOFT, background=GREEN, thickness=18,
                     borderwidth=0, lightcolor=GREEN, darkcolor=GREEN)

    # -- helpers de UI --
    def _boton(self, parent, texto, cmd, bg, hover, fg="white"):
        b = tk.Button(parent, text=texto, command=cmd, bg=bg, fg=fg,
                      activebackground=hover, activeforeground=fg,
                      relief="flat", bd=0, cursor="hand2",
                      font=("Segoe UI", 10, "bold"), padx=16, pady=10)
        b.bind("<Enter>", lambda e: b.config(bg=hover) if str(b["state"]) != "disabled" else None)
        b.bind("<Leave>", lambda e: b.config(bg=bg) if str(b["state"]) != "disabled" else None)
        return b

    def _tarjeta(self, parent, titulo):
        wrap = tk.Frame(parent, bg=BG)
        card = tk.Frame(wrap, bg=CARD, highlightbackground=BORDER,
                        highlightthickness=1, bd=0)
        card.pack(fill="both", expand=True)
        tk.Label(card, text=titulo, bg=CARD, fg=MUTED,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=14, pady=(10, 0))
        cuerpo = tk.Frame(card, bg=CARD)
        cuerpo.pack(fill="both", expand=True, padx=14, pady=10)
        return wrap, cuerpo

    def _construir_header(self):
        hdr = tk.Frame(self, bg=HEADER)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Panel de Gobernanza de Datos Maestros", bg=HEADER,
                 fg="white", font=("Segoe UI", 16, "bold")).pack(
            anchor="w", padx=20, pady=(16, 0))
        tk.Label(hdr, text="Depuracion de materiales  ·  Migracion SAP S/4HANA",
                 bg=HEADER, fg=HEADER_SUB, font=("Segoe UI", 10)).pack(
            anchor="w", padx=20, pady=(2, 16))

    def _construir_operacion(self, parent):
        wrap, cuerpo = self._tarjeta(parent, "OPERACION")
        wrap.pack(fill="x", pady=(0, 12))

        self.btn_foto = self._boton(cuerpo, "▶   Correr nueva corrida (foto)",
                                    self.correr_foto, GREEN, GREEN_H)
        self.btn_foto.grid(row=0, column=0, sticky="ew", padx=(0, 8))
        tk.Label(cuerpo, text="Ejecuta el query depurador y guarda una foto nueva (ST + RSS).",
                 bg=CARD, fg=MUTED, font=("Segoe UI", 8)).grid(row=1, column=0, sticky="w", pady=(4, 0))

        self.btn_excels = self._boton(cuerpo, "▶   Generar Excels de captura",
                                      self.generar_excels, BLUE, BLUE_H)
        self.btn_excels.grid(row=0, column=1, sticky="ew", padx=(8, 0))
        tk.Label(cuerpo, text="Crea los archivos por figura (productivo / no productivo).",
                 bg=CARD, fg=MUTED, font=("Segoe UI", 8)).grid(row=1, column=1, sticky="w", pady=(4, 0))

        cuerpo.columnconfigure(0, weight=1)
        cuerpo.columnconfigure(1, weight=1)

    def _construir_monitoreo(self, parent):
        wrap, cuerpo = self._tarjeta(parent, "AVANCE Y BUCKETS")
        wrap.pack(fill="both", expand=True, pady=(0, 12))

        # fila superior: % avance + barra + acciones
        top = tk.Frame(cuerpo, bg=CARD)
        top.pack(fill="x")

        self.lbl_pct = tk.Label(top, text="0.0%", bg=CARD, fg=GREEN,
                                font=("Segoe UI", 26, "bold"))
        self.lbl_pct.pack(side="left")

        medio = tk.Frame(top, bg=CARD)
        medio.pack(side="left", fill="x", expand=True, padx=16)
        tk.Label(medio, text="Avance global de decisiones", bg=CARD, fg=MUTED,
                 font=("Segoe UI", 9)).pack(anchor="w")
        self.pb = ttk.Progressbar(medio, style="Avance.Horizontal.TProgressbar",
                                  maximum=100, value=0)
        self.pb.pack(fill="x", pady=(4, 2))
        self.lbl_resumen = tk.Label(medio, text="", bg=CARD, fg=TEXT,
                                    font=("Segoe UI", 9))
        self.lbl_resumen.pack(anchor="w")

        acc = tk.Frame(top, bg=CARD)
        acc.pack(side="right")
        self._boton(acc, "Actualizar", self.actualizar, SOFT, SOFT_H, fg=BLUE).pack(
            side="left", padx=(0, 6))
        self._boton(acc, "Exportar reporte", self.exportar_reporte, SOFT, SOFT_H, fg=BLUE).pack(
            side="left")

        # tabla
        cols = ("Dominio", "Figura", "Total", "Decididos", "Pendientes",
                "% Avance", "Migrar", "Descartar", "Enriquecer")
        self.tabla = ttk.Treeview(cuerpo, columns=cols, show="headings", height=7)
        for c in cols:
            self.tabla.heading(c, text=c)
            self.tabla.column(c, width=96, anchor="center")
        self.tabla.column("Dominio", width=74)
        self.tabla.column("Figura", width=128)
        self.tabla.tag_configure("odd", background=STRIPE)
        self.tabla.tag_configure("even", background=CARD)
        self.tabla.pack(fill="both", expand=True, pady=(12, 0))

    def _construir_registro(self, parent):
        wrap, cuerpo = self._tarjeta(parent, "REGISTRO")
        wrap.pack(fill="both", expand=True)
        self.log = scrolledtext.ScrolledText(cuerpo, height=7, state="disabled",
                                             font=("Consolas", 9), bg="#F7F9FB",
                                             fg=TEXT, relief="flat", borderwidth=0)
        self.log.pack(fill="both", expand=True)

    # -- utilidades --
    def _log(self, msg):
        self.log.config(state="normal")
        self.log.insert("end", msg + "\n")
        self.log.see("end")
        self.log.config(state="disabled")

    def _set_ocupado(self, ocupado):
        self.ocupado = ocupado
        estado = "disabled" if ocupado else "normal"
        self.btn_foto.config(state=estado)
        self.btn_excels.config(state=estado)

    def _correr_script(self, args, titulo, al_terminar=None):
        if self.ocupado:
            return
        self._set_ocupado(True)
        self._log(f"\n=== {titulo} ===")

        def worker():
            try:
                proc = subprocess.Popen(
                    [PY, *args], cwd=str(PROJECT_ROOT),
                    stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                    text=True, encoding="utf-8", errors="replace",
                )
                for line in proc.stdout:
                    self.after(0, self._log, "   " + line.rstrip())
                proc.wait()
                self.after(0, self._log, f"=== {titulo}: terminado (codigo {proc.returncode}) ===")
            except Exception as e:  # noqa: BLE001
                self.after(0, self._log, f"ERROR: {e}")
            finally:
                self.after(0, self._set_ocupado, False)
                if al_terminar:
                    self.after(0, al_terminar)

        threading.Thread(target=worker, daemon=True).start()

    # -- acciones --
    def correr_foto(self):
        if not messagebox.askyesno(
            "Nueva corrida",
            "Esto ejecuta el query depurador y genera una FOTO nueva (ST y RSS).\n\n"
            "Despues conviene 'Generar Excels de captura'. ¿Continuar?",
        ):
            return
        self._correr_script([SCRIPT_SNAPSHOT], "Correr nueva corrida (foto)",
                            al_terminar=self.actualizar)

    def generar_excels(self):
        self._log("\n=== Generar Excels de captura (ST y RSS) ===")

        def paso2():
            self._correr_script([SCRIPT_EXPORT, "RSS"], "Generar Excels (RSS)")

        self._correr_script([SCRIPT_EXPORT, "ST"], "Generar Excels (ST)", al_terminar=paso2)

    def actualizar(self):
        try:
            with get_connection() as con:
                cur = con.cursor()
                avance = fetch_avance(cur)
                salio = fetch_salio(cur)
            self.tabla.delete(*self.tabla.get_children())
            tot = dec = 0
            for i, (dom, fig, t, de, p, pc, mi, ds, en) in enumerate(avance):
                self.tabla.insert("", "end",
                                  tags=("odd" if i % 2 else "even",),
                                  values=(dom, fig, f"{t:,}", f"{de:,}", f"{p:,}",
                                          f"{pc:.1f}%", f"{mi:,}", f"{ds:,}", f"{en:,}"))
                tot += t
                dec += de
            pct = (100.0 * dec / tot) if tot else 0.0
            n_salio = sum(n for _, n in salio)
            self.lbl_pct.config(text=f"{pct:.1f}%")
            self.pb.config(value=pct)
            self.lbl_resumen.config(
                text=f"{dec:,} de {tot:,} materiales decididos   ·   SALIO: {n_salio}")
            self._log("Monitor actualizado.")
        except Exception as e:  # noqa: BLE001
            self._log(f"ERROR al actualizar el monitor: {e}")

    def exportar_reporte(self):
        try:
            marca = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
            ruta = CARPETA_ENTREGABLES / f"reporte_avance_{marca}.xlsx"
            generar_reporte(ruta)
            self._log(f"Reporte generado: {ruta}")
            messagebox.showinfo("Reporte generado",
                                f"Se genero el reporte de avance:\n\n{ruta}")
        except Exception as e:  # noqa: BLE001
            self._log(f"ERROR al generar el reporte: {e}")
            messagebox.showerror("Error", f"No se pudo generar el reporte:\n\n{e}")


if __name__ == "__main__":
    Panel().mainloop()
