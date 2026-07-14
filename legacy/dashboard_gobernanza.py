"""Dashboard de Gobernanza de Datos Maestros  (Streamlit).

Herramienta del EQUIPO DE DATOS MAESTROS. Vista de dashboard web (corre local en el
navegador) con diseño a medida (tarjetas + barras medidor), operaciones y reporte.

  - Operaciones (barra lateral): correr nueva foto, generar Excels de captura.
  - KPIs, avance global, resumen visual (medidores), avance por figura, tabla.
  - Descarga del reporte en Excel.

Lanzar:
    doble clic en "Dashboard Gobernanza.bat"
    o:  python -m streamlit run dashboard_gobernanza.py
"""

from __future__ import annotations

import datetime as dt
import io
import subprocess
import sys

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from connection import PROJECT_ROOT, get_connection

SCRIPT_SNAPSHOT = "05_snapshot.py"
SCRIPT_EXPORT = "08_export_decisiones.py"

# Paleta (alineada al diseño aprobado)
BLUE, BLUE_SOFT, GREEN, RED, AMBER = "#0d6efd", "#7aa7ff", "#1f9d55", "#d64545", "#e0a100"
SLATE, GRAY, INK, MUTED = "#94a3b8", "#c3c9d2", "#1f2733", "#6b7684"


# ----------------------------------------------------------------------------
# Datos
# ----------------------------------------------------------------------------
@st.cache_data(ttl=30, show_spinner=False)
def cargar_datos():
    with get_connection() as con:
        cur = con.cursor()
        cur.execute(
            """SELECT PT_DOMAIN,
                      CASE WHEN UPPER(PT_PROD_LINE)='REF' THEN 'NO_PRODUCTIVO' ELSE 'PRODUCTIVO' END AS FIGURA,
                      COUNT(*) AS TOTAL,
                      SUM(CASE WHEN DECISION_PREVIA IN ('MIGRAR','DESCARTAR') THEN 1 ELSE 0 END) AS DECIDIDOS,
                      SUM(CASE WHEN DECISION_PREVIA='MIGRAR'     THEN 1 ELSE 0 END) AS MIGRAR,
                      SUM(CASE WHEN DECISION_PREVIA='DESCARTAR'  THEN 1 ELSE 0 END) AS DESCARTAR
                 FROM V_GD_MERGE_MATERIALES
                GROUP BY PT_DOMAIN,
                         CASE WHEN UPPER(PT_PROD_LINE)='REF' THEN 'NO_PRODUCTIVO' ELSE 'PRODUCTIVO' END
                ORDER BY PT_DOMAIN, FIGURA"""
        )
        av = pd.DataFrame(cur.fetchall(),
                          columns=["Dominio", "Figura", "Total", "Decididos",
                                   "Migrar", "Descartar"])
        av["Pendientes"] = av["Total"] - av["Decididos"]
        av["% Avance"] = (100.0 * av["Decididos"] / av["Total"]).round(1).fillna(0)

        cur.execute(
            """SELECT PT_DOMAIN, BUCKET, COUNT(*) FROM V_GD_MERGE_MATERIALES
                GROUP BY PT_DOMAIN, BUCKET ORDER BY PT_DOMAIN, BUCKET"""
        )
        bk = pd.DataFrame(cur.fetchall(), columns=["Dominio", "Bucket", "Materiales"])

        cur.execute(
            """SELECT ATENCION, COUNT(*) FROM V_GD_MERGE_SALIO
                GROUP BY ATENCION ORDER BY ATENCION"""
        )
        sl = pd.DataFrame(cur.fetchall(), columns=["Atencion", "Materiales"])

    tot = int(av["Total"].sum())
    dec = int(av["Decididos"].sum())
    return {"avance": av, "buckets": bk, "salio": sl, "tot": tot, "dec": dec,
            "pend": tot - dec, "pct": (100.0 * dec / tot) if tot else 0.0,
            "n_salio": int(sl["Materiales"].sum()) if not sl.empty else 0,
            "estados": {"Migrar": int(av["Migrar"].sum()),
                        "Descartar": int(av["Descartar"].sum())}}


def reporte_excel_bytes(d) -> bytes:
    wb = Workbook()
    hf, hfill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1F4E78")

    def escribe(ws, df):
        for j, h in enumerate(df.columns, start=1):
            c = ws.cell(row=1, column=j, value=str(h))
            c.font, c.fill = hf, hfill
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[chr(64 + j)].width = max(12, len(str(h)) + 2)
        for i, row in enumerate(df.itertuples(index=False), start=2):
            for j, v in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=v)

    ws1 = wb.active
    ws1.title = "Avance"
    escribe(ws1, d["avance"][["Dominio", "Figura", "Total", "Decididos",
                              "Pendientes", "% Avance", "Migrar", "Descartar"]])
    escribe(wb.create_sheet("Buckets"), d["buckets"])
    escribe(wb.create_sheet("SALIO"), d["salio"] if not d["salio"].empty
            else pd.DataFrame([{"Atencion": "(ninguno)", "Materiales": 0}]))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def correr_script(args):
    p = subprocess.run([sys.executable, *args], cwd=str(PROJECT_ROOT),
                       capture_output=True, text=True, encoding="utf-8", errors="replace")
    return p.returncode, (p.stdout or "") + (p.stderr or "")


# ----------------------------------------------------------------------------
# Componentes HTML
# ----------------------------------------------------------------------------
def fmt(n):
    return f"{n:,}"


def medidor(segmentos, total):
    """segmentos = [(label, color, valor)] -> (html_barra, html_leyenda)."""
    barra, leyenda = "", ""
    for label, color, val in segmentos:
        pct = (100.0 * val / total) if total else 0.0
        barra += f'<span style="width:{pct:.2f}%;background:{color}"></span>'
        leyenda += (f'<div class="lr"><span class="dot" style="background:{color}"></span>'
                    f'<span class="ll">{label}</span><span class="lv">{fmt(val)}</span>'
                    f'<span class="lp">{pct:.1f}%</span></div>')
    return f'<div class="meter">{barra}</div>', f'<div class="legend">{leyenda}</div>'


def tarjeta_medidor(titulo, big, big_color, big_lbl, segmentos, total):
    barra, leyenda = medidor(segmentos, total)
    return (f'<div class="card"><div class="ct">{titulo}</div>'
            f'<div class="big"><span class="bigval" style="color:{big_color}">{big}</span>'
            f'<span class="biglbl">{big_lbl}</span></div>{barra}{leyenda}</div>')


def barras_horizontales(titulo, subtitulo, filas):
    """filas = [(label, valor, pct, color, sufijo)]."""
    cuerpo = ""
    for label, valor, pct, color, sufijo in filas:
        cuerpo += (
            f'<div class="statrow"><div class="statrow-top"><span class="sl">{label}</span>'
            f'<span class="sv">{valor}<span class="sp">{sufijo}</span></span></div>'
            f'<div class="track"><span style="width:{max(pct,0):.1f}%;background:{color}"></span></div></div>')
    sub = f'<span class="ct-sub">{subtitulo}</span>' if subtitulo else ""
    return f'<div class="card"><div class="ct">{titulo}{sub}</div>{cuerpo}</div>'


CSS = f"""
<style>
  .stApp {{ background:#f1f3f6; }}
  #MainMenu, footer, header {{ visibility:hidden; }}
  .block-container {{ padding-top:1.4rem; padding-bottom:2rem; max-width:1280px; }}
  .gd-banner {{ background:#fff; border:1px solid #e6e9ef; border-left:5px solid {BLUE};
    border-radius:14px; padding:20px 26px; display:flex; justify-content:space-between;
    align-items:center; gap:16px; box-shadow:0 1px 3px rgba(16,24,40,.05); }}
  .gd-banner h1 {{ margin:0; font-size:22px; color:{INK}; font-weight:800; }}
  .gd-banner p {{ margin:5px 0 0; color:{MUTED}; font-size:13px; }}
  .gd-pill {{ background:#eaf2ff; color:{BLUE}; border:1px solid #cfe0ff; font-size:12px;
    padding:7px 14px; border-radius:999px; white-space:nowrap; font-weight:600; }}
  .gd-sec {{ font-size:12px; font-weight:700; color:{MUTED}; text-transform:uppercase;
    letter-spacing:.6px; margin:22px 2px 10px; }}
  .kpis {{ display:grid; grid-template-columns:repeat(5,1fr); gap:14px; }}
  .kpi {{ background:#fff; border:1px solid #e6e9ef; border-radius:14px; padding:15px 18px;
    box-shadow:0 1px 3px rgba(16,24,40,.05); position:relative; overflow:hidden; }}
  .kpi::before {{ content:""; position:absolute; left:0; top:0; bottom:0; width:4px; background:var(--a,{BLUE}); }}
  .kpi.hi {{ border:2px solid {BLUE}; }}
  .kpi .lbl {{ font-size:12px; color:{MUTED}; font-weight:700; }}
  .kpi .val {{ font-size:29px; font-weight:800; color:{INK}; margin-top:5px; font-variant-numeric:tabular-nums; }}
  .kpi.hi .val {{ color:{BLUE}; }}
  .kpi .sub {{ font-size:12px; color:{MUTED}; margin-top:2px; }}
  .card {{ background:#fff; border:1px solid #e6e9ef; border-radius:14px; padding:18px 20px;
    box-shadow:0 1px 3px rgba(16,24,40,.05); height:100%; }}
  .grid3 {{ display:grid; grid-template-columns:repeat(3,1fr); gap:14px; align-items:stretch; }}
  .grid2 {{ display:grid; grid-template-columns:repeat(2,1fr); gap:14px; align-items:stretch; margin-top:14px; }}
  @media (max-width:900px){{ .grid3,.grid2{{ grid-template-columns:1fr; }} .kpis{{ grid-template-columns:repeat(2,1fr); }} }}
  .ct {{ font-size:14px; font-weight:700; color:{INK}; margin-bottom:12px;
    display:flex; justify-content:space-between; align-items:baseline; }}
  .ct-sub {{ font-size:12px; font-weight:500; color:{MUTED}; }}
  .big {{ display:flex; align-items:baseline; gap:8px; margin-bottom:12px; }}
  .bigval {{ font-size:32px; font-weight:800; font-variant-numeric:tabular-nums; }}
  .biglbl {{ font-size:12px; color:{MUTED}; }}
  .meter {{ height:12px; background:#e9edf2; border-radius:999px; overflow:hidden; display:flex; }}
  .meter > span {{ display:block; height:100%; }}
  .legend {{ display:flex; flex-direction:column; gap:7px; margin-top:12px; font-size:13px; }}
  .lr {{ display:flex; align-items:center; gap:8px; }}
  .dot {{ width:11px; height:11px; border-radius:3px; flex:0 0 auto; display:inline-block; }}
  .ll {{ color:{INK}; }} .lv {{ margin-left:auto; font-weight:700; color:{INK}; font-variant-numeric:tabular-nums; }}
  .lp {{ color:{MUTED}; width:52px; text-align:right; font-variant-numeric:tabular-nums; }}
  .statrow {{ margin:11px 0; }}
  .statrow-top {{ display:flex; justify-content:space-between; font-size:13px; margin-bottom:5px; }}
  .statrow-top .sl {{ color:{INK}; font-weight:600; }}
  .sv {{ font-weight:700; color:{INK}; font-variant-numeric:tabular-nums; }}
  .sp {{ color:{MUTED}; font-weight:500; margin-left:8px; }}
  .track {{ height:12px; background:#e9edf2; border-radius:999px; overflow:hidden; }}
  .track > span {{ display:block; height:100%; }}
  .prog-top {{ display:flex; justify-content:space-between; font-size:13px; margin-bottom:8px; }}
  .prog-top .pv {{ font-weight:700; color:{INK}; font-variant-numeric:tabular-nums; }}
</style>
"""


# ----------------------------------------------------------------------------
# Página
# ----------------------------------------------------------------------------
st.set_page_config(page_title="Gobernanza de Datos Maestros", page_icon="🧭", layout="wide")
st.markdown(CSS, unsafe_allow_html=True)

# -- Barra lateral --
with st.sidebar:
    st.header("Operaciones")
    st.caption("Equipo de datos maestros")
    if st.button("▶  Correr nueva corrida (foto)", width="stretch"):
        with st.spinner("Ejecutando el query depurador y guardando la foto..."):
            rc, out = correr_script([SCRIPT_SNAPSHOT])
        st.session_state["log"] = out
        st.cache_data.clear()
        (st.success if rc == 0 else st.error)(
            "Foto generada." if rc == 0 else "La corrida terminó con errores.")
    if st.button("▶  Generar Excels de captura", width="stretch"):
        with st.spinner("Generando Excels por figura (ST y RSS)..."):
            rc1, o1 = correr_script([SCRIPT_EXPORT, "ST"])
            rc2, o2 = correr_script([SCRIPT_EXPORT, "RSS"])
        st.session_state["log"] = o1 + "\n" + o2
        (st.success if rc1 == rc2 == 0 else st.error)(
            "Excels generados en 'entregables'." if rc1 == rc2 == 0 else "Terminó con errores.")
    st.divider()
    if st.button("🔄  Actualizar datos", width="stretch"):
        st.cache_data.clear()
        st.rerun()
    if "log" in st.session_state:
        with st.expander("Ver registro de la última operación"):
            st.code(st.session_state["log"] or "(sin salida)")

# -- Datos --
try:
    d = cargar_datos()
except Exception as e:  # noqa: BLE001
    st.error(f"No se pudo conectar a la base de datos:\n\n{e}")
    st.stop()

ahora = dt.datetime.now().strftime("%Y-%m-%d %H:%M")

# -- Banner --
st.markdown(
    f'<div class="gd-banner"><div><h1>🧭 Gobernanza de Datos Maestros</h1>'
    f'<p>Depuración de materiales · Migración SAP S/4HANA</p></div>'
    f'<span class="gd-pill">Actualizado {ahora}</span></div>',
    unsafe_allow_html=True,
)

# -- KPIs --
st.markdown('<div class="gd-sec">Indicadores</div>', unsafe_allow_html=True)
st.markdown(
    f'<div class="kpis">'
    f'<div class="kpi" style="--a:{BLUE}"><div class="lbl">Candidatos</div><div class="val">{fmt(d["tot"])}</div><div class="sub">último snapshot</div></div>'
    f'<div class="kpi" style="--a:{GREEN}"><div class="lbl">Decididos</div><div class="val">{fmt(d["dec"])}</div><div class="sub">{d["pct"]:.1f}% del total</div></div>'
    f'<div class="kpi" style="--a:{AMBER}"><div class="lbl">Pendientes</div><div class="val">{fmt(d["pend"])}</div><div class="sub">por decidir</div></div>'
    f'<div class="kpi hi"><div class="lbl">% Avance</div><div class="val">{d["pct"]:.1f}%</div><div class="sub">avance global</div></div>'
    f'<div class="kpi" style="--a:{SLATE}"><div class="lbl">Salió</div><div class="val">{fmt(d["n_salio"])}</div><div class="sub">en vigilancia</div></div>'
    f'</div>',
    unsafe_allow_html=True,
)

# -- Avance global (barra) --
st.markdown(
    f'<div class="card" style="margin-top:14px"><div class="prog-top">'
    f'<span style="color:{MUTED}">Avance global de decisiones</span>'
    f'<span class="pv">{fmt(d["dec"])} / {fmt(d["tot"])}</span></div>'
    f'<div class="track"><span style="width:{d["pct"]:.1f}%;background:{BLUE}"></span></div></div>',
    unsafe_allow_html=True,
)

# -- Resumen visual (3 medidores) --
st.markdown('<div class="gd-sec">Resumen visual</div>', unsafe_allow_html=True)
bk_tot = d["buckets"].groupby("Bucket")["Materiales"].sum().to_dict()
fig_tot = d["avance"].groupby("Figura")["Total"].sum().to_dict()

tarj = (
    tarjeta_medidor("Avance global", f'{d["pct"]:.1f}%', BLUE, "decidido",
        [("Decididos", BLUE, d["dec"]), ("Pendientes", GRAY, d["pend"])], d["tot"])
    + tarjeta_medidor("Por bucket", fmt(d["tot"]), INK, "materiales",
        [("Vigente", BLUE, bk_tot.get("VIGENTE", 0)),
         ("Por decidir", BLUE_SOFT, bk_tot.get("POR_DECIDIR", 0)),
         ("Re-revisar", SLATE, bk_tot.get("RE_REVISAR", 0))], d["tot"])
    + tarjeta_medidor("Por figura", str(len(fig_tot)), INK, "figuras",
        [("Productivo", GREEN, fig_tot.get("PRODUCTIVO", 0)),
         ("No productivo", GRAY, fig_tot.get("NO_PRODUCTIVO", 0))], d["tot"])
)
st.markdown(f'<div class="grid3">{tarj}</div>', unsafe_allow_html=True)

# -- Decisiones por estado + Avance por figura --
dec = max(d["dec"], 1)
est = d["estados"]
card_est = barras_horizontales(
    "Decisiones por estado", f'{fmt(d["dec"])} decididos',
    [("Migrar", fmt(est["Migrar"]), 100.0 * est["Migrar"] / dec, GREEN, f'{100.0*est["Migrar"]/dec:.1f}%'),
     ("Descartar", fmt(est["Descartar"]), 100.0 * est["Descartar"] / dec, RED, f'{100.0*est["Descartar"]/dec:.1f}%')])
filas = []
for _, r in d["avance"].iterrows():
    filas.append((f'{r["Dominio"]} · {r["Figura"].title().replace("_", " ")}',
                  f'{fmt(int(r["Decididos"]))} / {fmt(int(r["Total"]))}',
                  float(r["% Avance"]), BLUE, f'{r["% Avance"]:.0f}%'))
card_fig = barras_horizontales("Avance por figura", "", filas)
st.markdown(f'<div class="grid2">{card_est}{card_fig}</div>', unsafe_allow_html=True)

# -- Tabla + descarga --
st.markdown('<div class="gd-sec">Detalle de avance</div>', unsafe_allow_html=True)
st.dataframe(
    d["avance"][["Dominio", "Figura", "Total", "Decididos", "Pendientes",
                 "% Avance", "Migrar", "Descartar"]],
    width="stretch", hide_index=True,
)
st.download_button(
    "📄  Descargar reporte de avance (Excel)",
    data=reporte_excel_bytes(d),
    file_name="reporte_avance_gobernanza.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
