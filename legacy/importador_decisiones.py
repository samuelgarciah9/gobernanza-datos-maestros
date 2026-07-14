"""Importador de Decisiones  (Streamlit)  -  App para las FIGURAS que analizan.

Cada figura (productivo / no productivo) sube SU Excel de captura y las decisiones
se vacían a GD_DECISIONES_MATERIALES de forma autónoma. UPSERT re-ejecutable.

Mismas convenciones de diseño que el dashboard (banner, tarjetas, paleta, azul).

Lanzar:
    doble clic en "Importador de decisiones.bat"
    o:  python -m streamlit run importador_decisiones.py
"""

from __future__ import annotations

import io
import json

import streamlit as st
from openpyxl import load_workbook

from connection import PROJECT_ROOT, get_connection

TABLA = "GD_DECISIONES_MATERIALES"
ESTADOS_VALIDOS = {"MIGRAR", "DESCARTAR", "PENDIENTE"}
ROLES = ["COMPRAS", "INVENTARIOS", "INGENIERIA"]
CONFIG = PROJECT_ROOT / ".importador_config.json"

HOJA = "Decisiones"
COL_MAT, COL_DOM = "NUMERO_PRODUCTO_ANTIGUO", "PT_DOMAIN"
COL_ESTADO, COL_RAZON, COL_COMENT = "ESTADO", "RAZON", "COMENTARIO"
COL_RUN, COL_HASH = "RUN_ID_AL_DECIDIR", "HASH_AL_DECIDIR"

BLUE, GREEN, RED, AMBER, INK, MUTED = "#0d6efd", "#1f9d55", "#d64545", "#e0a100", "#1f2733", "#6b7684"


# ----------------------------------------------------------------------------
# Lógica (igual que la versión previa; lee de un archivo en memoria)
# ----------------------------------------------------------------------------
def leer_decisiones(archivo):
    wb = load_workbook(archivo, data_only=True)
    if HOJA not in wb.sheetnames:
        raise ValueError(f"El archivo no tiene la hoja '{HOJA}'. ¿Es el Excel de captura correcto?")
    ws = wb[HOJA]
    hdr = {}
    for idx, cell in enumerate(ws[1]):
        if cell.value is not None:
            hdr[str(cell.value).strip()] = idx
    faltan = [c for c in (COL_MAT, COL_ESTADO, COL_RAZON, COL_RUN, COL_HASH) if c not in hdr]
    if faltan:
        raise ValueError(f"Al Excel le faltan columnas requeridas: {', '.join(faltan)}")

    tiene_dom, tiene_coment = COL_DOM in hdr, COL_COMENT in hdr

    def val(row, col):
        v = row[hdr[col]].value if col in hdr else None
        if v is None:
            return None
        v = str(v).strip()
        return v or None

    decisiones, errores, sin_decidir = [], [], 0
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if all(c.value in (None, "") for c in row):
            continue
        estado = val(row, COL_ESTADO)
        if not estado:
            sin_decidir += 1
            continue
        estado = estado.upper()
        mat, razon = val(row, COL_MAT), val(row, COL_RAZON)
        coment = val(row, COL_COMENT) if tiene_coment else None
        run, hsh = val(row, COL_RUN), val(row, COL_HASH)
        dom = val(row, COL_DOM) if tiene_dom else None
        if not dom and run:
            partes = run.split("_")
            dom = partes[1].upper() if len(partes) >= 2 else None
        dom = dom.upper() if dom else None

        etq = f"Fila {i}" + (f" (material {mat})" if mat else "")
        if estado not in ESTADOS_VALIDOS:
            errores.append(f"{etq}: ESTADO inválido '{estado}'.")
            continue
        if estado != "PENDIENTE" and not razon:
            errores.append(f"{etq}: falta RAZÓN (obligatoria si ESTADO no es PENDIENTE).")
            continue
        if not (mat and dom and run and hsh):
            errores.append(f"{etq}: faltan datos de control (material/dominio/run/hash).")
            continue
        decisiones.append({"mat": mat, "dom": dom, "estado": estado, "razon": razon,
                           "coment": coment, "run": run, "hash": hsh})
    return decisiones, errores, sin_decidir


MERGE_SQL = f"""
MERGE INTO {TABLA} d
USING (SELECT :mat AS NUMERO_PRODUCTO_ANTIGUO, :dom AS PT_DOMAIN FROM dual) s
   ON (d.NUMERO_PRODUCTO_ANTIGUO = s.NUMERO_PRODUCTO_ANTIGUO AND d.PT_DOMAIN = s.PT_DOMAIN)
WHEN MATCHED THEN UPDATE SET
     ESTADO = :estado, RAZON = :razon, COMENTARIO = :coment,
     DECIDIDO_POR = :quien, ROL = :rol, FECHA_DECISION = SYSDATE,
     HASH_AL_DECIDIR = :hash, RUN_ID_AL_DECIDIR = :run
WHEN NOT MATCHED THEN INSERT
     (NUMERO_PRODUCTO_ANTIGUO, PT_DOMAIN, ESTADO, RAZON, COMENTARIO,
      DECIDIDO_POR, ROL, FECHA_DECISION, HASH_AL_DECIDIR, RUN_ID_AL_DECIDIR)
     VALUES (:mat, :dom, :estado, :razon, :coment, :quien, :rol, SYSDATE, :hash, :run)
"""


def guardar_decisiones(decisiones, quien, rol):
    with get_connection() as con:
        cur = con.cursor()
        cur.execute(f"SELECT NUMERO_PRODUCTO_ANTIGUO, PT_DOMAIN FROM {TABLA}")
        existentes = {(m, dm) for m, dm in cur.fetchall()}
        nuevas = sum(1 for x in decisiones if (x["mat"], x["dom"]) not in existentes)
        binds = [dict(x, quien=quien[:60], rol=rol) for x in decisiones]
        cur.executemany(MERGE_SQL, binds)
        con.commit()
    return nuevas, len(decisiones) - nuevas


def cargar_config():
    try:
        return json.loads(CONFIG.read_text(encoding="utf-8"))
    except Exception:
        return {}


def guardar_config(nombre, rol):
    try:
        CONFIG.write_text(json.dumps({"nombre": nombre, "rol": rol}), encoding="utf-8")
    except Exception:
        pass


# ----------------------------------------------------------------------------
# Interfaz
# ----------------------------------------------------------------------------
st.set_page_config(page_title="Importar Decisiones", page_icon="📥", layout="centered")

st.markdown(f"""
<style>
  .stApp {{ background:#f1f3f6; }}
  #MainMenu, footer, header {{ visibility:hidden; }}
  .block-container {{ padding-top:1.4rem; max-width:820px; }}
  .gd-banner {{ background:#fff; border:1px solid #e6e9ef; border-left:5px solid {BLUE};
    border-radius:14px; padding:20px 26px; box-shadow:0 1px 3px rgba(16,24,40,.05); margin-bottom:8px; }}
  .gd-banner h1 {{ margin:0; font-size:22px; color:{INK}; font-weight:800; }}
  .gd-banner p {{ margin:5px 0 0; color:{MUTED}; font-size:13px; }}
  .gd-sec {{ font-size:12px; font-weight:700; color:{MUTED}; text-transform:uppercase;
    letter-spacing:.6px; margin:20px 2px 8px; }}
  /* tarjetas de contenedores nativos */
  div[data-testid="stVerticalBlockBorderWrapper"] {{ background:#fff; border:1px solid #e6e9ef !important;
    border-radius:14px; box-shadow:0 1px 3px rgba(16,24,40,.05); }}
  /* botón primario a ancho completo */
  div.stButton > button {{ font-weight:700; border-radius:10px; }}
  /* KPIs de resultado */
  .kpis {{ display:grid; grid-template-columns:repeat(4,1fr); gap:12px; margin-top:6px; }}
  .kpi {{ background:#fff; border:1px solid #e6e9ef; border-radius:14px; padding:14px 16px;
    box-shadow:0 1px 3px rgba(16,24,40,.05); position:relative; overflow:hidden; }}
  .kpi::before {{ content:""; position:absolute; left:0; top:0; bottom:0; width:4px; background:var(--a,{BLUE}); }}
  .kpi .lbl {{ font-size:12px; color:{MUTED}; font-weight:700; }}
  .kpi .val {{ font-size:26px; font-weight:800; color:{INK}; margin-top:4px; font-variant-numeric:tabular-nums; }}
  @media(max-width:680px){{ .kpis{{ grid-template-columns:repeat(2,1fr); }} }}
</style>
""", unsafe_allow_html=True)

st.markdown(
    '<div class="gd-banner"><h1>📥 Importar Decisiones</h1>'
    '<p>Carga tu Excel de captura a la base · Gobernanza de Datos Maestros</p></div>',
    unsafe_allow_html=True,
)

cfg = cargar_config()
st.session_state.setdefault("nombre", cfg.get("nombre", ""))
st.session_state.setdefault("rol", cfg.get("rol", ROLES[0]))

st.markdown('<div class="gd-sec">Revisor</div>', unsafe_allow_html=True)
with st.container(border=True):
    col_a, col_b = st.columns([2, 1])
    with col_a:
        st.text_input("Nombre", key="nombre", placeholder="Tu nombre completo")
    with col_b:
        st.selectbox("Rol", ROLES, key="rol")

st.markdown('<div class="gd-sec">Archivo de decisiones</div>', unsafe_allow_html=True)
with st.container(border=True):
    archivo = st.file_uploader("Selecciona tu Excel de captura (.xlsx)", type=["xlsx"])

cargar = st.button("📤  Cargar decisiones", type="primary", width="stretch")

if cargar:
    nombre = (st.session_state.get("nombre") or "").strip()
    rol = st.session_state.get("rol")
    if not nombre:
        st.warning("Escribe tu nombre antes de cargar.")
    elif not archivo:
        st.warning("Selecciona tu Excel de decisiones.")
    else:
        try:
            decisiones, errores, sin_decidir = leer_decisiones(io.BytesIO(archivo.getvalue()))
            if not decisiones:
                st.info("No hay decisiones válidas que cargar (todas las filas están sin ESTADO).")
                st.session_state["resultado"] = {"nuevas": 0, "actualizadas": 0,
                                                 "sin_decidir": sin_decidir, "errores": errores}
            else:
                with st.spinner("Guardando en la base de datos..."):
                    nuevas, actualizadas = guardar_decisiones(decisiones, nombre, rol)
                guardar_config(nombre, rol)
                st.session_state["resultado"] = {"nuevas": nuevas, "actualizadas": actualizadas,
                                                 "sin_decidir": sin_decidir, "errores": errores}
                st.success(f"Decisiones guardadas: {nuevas} nuevas, {actualizadas} actualizadas.")
        except Exception as e:  # noqa: BLE001
            st.error(f"No se pudo completar la carga:\n\n{e}")
            st.session_state.pop("resultado", None)

# -- Resultado --
r = st.session_state.get("resultado")
if r:
    st.markdown('<div class="gd-sec">Resultado de la última carga</div>', unsafe_allow_html=True)
    st.markdown(
        f'<div class="kpis">'
        f'<div class="kpi" style="--a:{GREEN}"><div class="lbl">Nuevas</div><div class="val">{r["nuevas"]:,}</div></div>'
        f'<div class="kpi" style="--a:{BLUE}"><div class="lbl">Actualizadas</div><div class="val">{r["actualizadas"]:,}</div></div>'
        f'<div class="kpi" style="--a:{AMBER}"><div class="lbl">Sin decidir</div><div class="val">{r["sin_decidir"]:,}</div></div>'
        f'<div class="kpi" style="--a:{RED}"><div class="lbl">Con problema</div><div class="val">{len(r["errores"]):,}</div></div>'
        f'</div>',
        unsafe_allow_html=True,
    )
    if r["errores"]:
        with st.expander(f"Ver {len(r['errores'])} fila(s) con problema (no se cargaron)"):
            st.write("\n".join(f"- {e}" for e in r["errores"][:200]))
